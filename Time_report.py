import pandas as pd  
import matplotlib.pyplot as plt
import seaborn as sns
import os
import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from matplotlib.backends.backend_pdf import PdfPages

sns.set(style="whitegrid")

def setup_paths():
    base_dir = "Time_report.xlsm"
    today = datetime.datetime.today().strftime('%Y%m%d')
    return {
        'base_dir': base_dir,
        'template_file': os.path.join(base_dir, "Time_report.xlsm"),
        'output_file': os.path.join(base_dir, f"Time_report_{today}.xlsx"),
        'chart_dir': os.path.join(base_dir, "charts_temp"),
        'chart_project_dir': os.path.join(base_dir, "charts_project"),
        'pdf_report': os.path.join(base_dir, "charts_report.pdf")
    }

def read_configs(path_dict):
    year_mode_df = pd.read_excel(path_dict['template_file'], sheet_name='Config_Year_Mode', engine='openpyxl')
    project_filter_df = pd.read_excel(path_dict['template_file'], sheet_name='Config_Project_Filter', engine='openpyxl')

    if 'Key' not in year_mode_df.columns or 'Value' not in year_mode_df.columns:
        raise ValueError("⚠️ 'Config_Year_Mode' must have 'Key' and 'Value' columns.")

    required_keys = ['mode', 'year', 'months']
    for key in required_keys:
        if year_mode_df[year_mode_df['Key'].str.lower() == key].empty:
            raise ValueError(f"⚠️ Missing required key: '{key}' in 'Config_Year_Mode'")

    mode = year_mode_df.loc[year_mode_df['Key'].str.lower() == 'mode', 'Value'].values[0].strip().lower()

    year_row = year_mode_df.loc[year_mode_df['Key'].str.lower() == 'year', 'Value']
    year = None
    if not year_row.empty and pd.notna(year_row.values[0]):
        try:
            year = int(year_row.values[0])
        except ValueError:
            pass

    months_row = year_mode_df.loc[year_mode_df['Key'].str.lower() == 'months', 'Value']
    months = []
    if not months_row.empty and pd.notna(months_row.values[0]):
        months_raw = months_row.values[0]
        months = [m.strip().capitalize() for m in str(months_raw).split(',') if m.strip()]

    print("⚙️ Loaded config:")
    print("  Mode:", mode)
    print("  Year:", year if year is not None else "All")
    print("  Months:", months)

    return {
        'mode': mode,
        'year': year,
        'months': months,
        'project_filter_df': project_filter_df
    }

def load_raw_data(path_dict):
    df = pd.read_excel(path_dict['template_file'], sheet_name='Raw Data', engine='openpyxl')
    df.rename(columns={'Team member': 'Employee', 'Hou': 'Hours'}, inplace=True)
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df['Year'] = df['Date'].dt.year
    df['MonthName'] = df['Date'].dt.month_name()
    df['Week'] = df['Date'].dt.isocalendar().week
    print(f"📥 Loaded raw data: {len(df)} rows")
    return df

def save_chart(fig, path):
    fig.tight_layout()
    fig.savefig(path)
    plt.close(fig)

def generate_general_charts(df, chart_dir):
    os.makedirs(chart_dir, exist_ok=True)

    fig, ax = plt.subplots(figsize=(10, 6))
    df.groupby('Project name')['Hours'].sum().sort_values().plot(kind='barh', ax=ax, color='skyblue')
    ax.set_title('Total Hours by Project')
    save_chart(fig, os.path.join(chart_dir, '1_project_hours.png'))

    fig, ax = plt.subplots(figsize=(10, 6))
    df.groupby('Workcentre')['Hours'].sum().sort_values().plot(kind='barh', ax=ax, color='orange')
    ax.set_title('Total Hours by Workcentre')
    save_chart(fig, os.path.join(chart_dir, '2_workcentre_hours.png'))

    fig, ax = plt.subplots(figsize=(10, 6))
    df.groupby(df['Date'].dt.to_period('M'))['Hours'].sum().plot(marker='o', ax=ax, color='green')
    ax.set_title('Monthly Trend')
    save_chart(fig, os.path.join(chart_dir, '3_monthly_trend.png'))

def generate_project_chart(df_proj, project_name, chart_project_dir):
    os.makedirs(chart_project_dir, exist_ok=True)
    fig, ax = plt.subplots(figsize=(10, 6))
    df_proj.groupby('Workcentre')['Hours'].sum().sort_values().plot(kind='barh', ax=ax, color='teal')
    ax.set_title(f'{project_name} - Hours by Workcentre')
    path = os.path.join(chart_project_dir, f"{project_name[:31]}.png")
    save_chart(fig, path)
    return path

def apply_filters(df, config):
    if config['year'] is not None:
        df_filtered = df[df['Year'] == config['year']]
    else:
        df_filtered = df.copy()
        print("📆 No year filter — using all years.")

    if config['months']:
        df_filtered = df_filtered[df_filtered['MonthName'].isin(config['months'])]
    else:
        print("📅 No month filter — using all months in the year.")

    df_filtered = df_filtered.merge(
        config['project_filter_df'][config['project_filter_df']['Include'].str.lower() == 'yes'],
        how='inner',
        left_on='Project name',
        right_on='Project Name'
    )
    print(f"🔎 Filtered data: {len(df_filtered)} rows")
    return df_filtered

def add_project_analysis_sheet(wb, df_project, project_name):
    ws = wb.create_sheet(title=project_name[:31])

    for r in dataframe_to_rows(df_project, index=False, header=True):
        ws.append(r)

    start_row = len(df_project) + 3
    ws.cell(row=start_row, column=1, value="Workcentre")
    ws.cell(row=start_row, column=2, value="Total Hours")

    workcentre_summary = df_project.groupby('Workcentre')['Hours'].sum().reset_index()
    for i, row in enumerate(workcentre_summary.itertuples(index=False), start=start_row + 1):
        ws.cell(row=i, column=1, value=row.Workcentre)
        ws.cell(row=i, column=2, value=row.Hours)

    chart = BarChart()
    chart.title = f"{project_name} - Hours by Workcentre"
    chart.y_axis.title = "Hours"
    chart.x_axis.title = "Workcentre"

    data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=i)
    cats_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=i)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, f"E{start_row}")

def export_all_charts_to_pdf(path_dict):
    chart_paths = []

    for chart_file in sorted(os.listdir(path_dict['chart_dir'])):
        full_path = os.path.join(path_dict['chart_dir'], chart_file)
        if os.path.isfile(full_path) and chart_file.endswith(".png"):
            chart_paths.append(full_path)

    for chart_file in sorted(os.listdir(path_dict['chart_project_dir'])):
        full_path = os.path.join(path_dict['chart_project_dir'], chart_file)
        if os.path.isfile(full_path) and chart_file.endswith(".png"):
            chart_paths.append(full_path)

    with PdfPages(path_dict['pdf_report']) as pdf:
        for chart_file in chart_paths:
            fig, ax = plt.subplots(figsize=(11.7, 8.3))
            img = plt.imread(chart_file)
            ax.imshow(img)
            ax.axis('off')
            pdf.savefig(fig)
            plt.close(fig)

    print(f"🧾 PDF charts report saved: {path_dict['pdf_report']}")

def export_report(df, config, path_dict):
    mode = config['mode']
    if mode == 'year':
        summary = df.groupby(['Year', 'Project name'])['Hours'].sum().reset_index()
    elif mode == 'month':
        summary = df.groupby(['Year', 'MonthName', 'Project name'])['Hours'].sum().reset_index()
    else:
        summary = df.groupby(['Year', 'Week', 'Project name'])['Hours'].sum().reset_index()

    generate_general_charts(df, path_dict['chart_dir'])

    with pd.ExcelWriter(path_dict['output_file'], engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Raw_Data', index=False)
        summary.to_excel(writer, sheet_name='Summary', index=False)

    wb = load_workbook(path_dict['output_file'])

    for project in df['Project name'].unique():
        df_proj = df[df['Project name'] == project]
        generate_project_chart(df_proj, project, path_dict['chart_project_dir'])
        add_project_analysis_sheet(wb, df_proj, project)

    ws = wb.create_sheet("Charts")
    row = 1
    col_offset = 7

    for folder in [path_dict['chart_dir'], path_dict['chart_project_dir']]:
        for chart_file in sorted(os.listdir(folder)):
            chart_path = os.path.join(folder, chart_file)
            if os.path.isfile(chart_path):
                try:
                    img = ExcelImage(chart_path)
                    img.width = 640
                    img.height = 360
                    ws.add_image(img, f"A{row}")

                    ws.cell(row=row, column=col_offset, value="Workcentre")
                    ws.cell(row=row, column=col_offset + 1, value="Hours")
                    sample_data = [("A", 10), ("B", 20), ("C", 15)]
                    for i, (wc, hrs) in enumerate(sample_data, start=row + 1):
                        ws.cell(row=i, column=col_offset, value=wc)
                        ws.cell(row=i, column=col_offset + 1, value=hrs)

                    chart = BarChart()
                    chart.title = f"Editable chart from {chart_file[:25]}"
                    chart.x_axis.title = "Workcentre"
                    chart.y_axis.title = "Hours"
                    data = Reference(ws, min_col=col_offset + 1, min_row=row, max_row=row + len(sample_data))
                    cats = Reference(ws, min_col=col_offset, min_row=row + 1, max_row=row + len(sample_data))
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                    ws.add_chart(chart, f"H{row}")

                    row += 20
                except Exception as e:
                    print(f"⚠️ Could not insert image/chart for {chart_file}: {e}")

    ws_config = wb.create_sheet("Config_Info")
    ws_config['A1'], ws_config['B1'] = "Mode", config['mode']
    ws_config['A2'], ws_config['B2'] = "Year", str(config['year']) if config['year'] is not None else "All"
    ws_config['A3'], ws_config['B3'] = "Months", ', '.join(config['months']) if config['months'] else 'All'
    ws_config['A4'], ws_config['B4'] = "Included Projects", ', '.join(
        config['project_filter_df'][config['project_filter_df']['Include'].str.lower() == 'yes']['Project Name']
    )

    wb.save(path_dict['output_file'])
    print(f"✅ Excel report saved: {path_dict['output_file']}")

    export_all_charts_to_pdf(path_dict)

def main():
    path_dict = setup_paths()
    if not os.path.exists(path_dict['template_file']):
        print(f"❌ Template file not found: {path_dict['template_file']}")
        return
    df_raw = load_raw_data(path_dict)
    config = read_configs(path_dict)
    df_filtered = apply_filters(df_raw, config)
    if df_filtered.empty:
        print("⚠️ No data after filtering. Please check your config.")
        return
    export_report(df_filtered, config, path_dict)

if __name__ == "__main__":
    main()
