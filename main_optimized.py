import pandas as pd
import datetime
import os
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils.dataframe import dataframe_to_rows
from fpdf import FPDF
from matplotlib import pyplot as plt
import tempfile
import re
import shutil

# --- Language Data ---
def initialize_language_data():
    return {
        "en": {
            "app_title": "Project Time Report Application",
            "upload_excel_file": "Upload Excel file (.xlsx)",
            "loading_data_spinner": "Loading and processing data...",
            "file_upload_success": "File uploaded successfully!",
            "file_upload_error": "Error loading file. Please check the format.",
            "upload_file_to_start": "Please upload an Excel file to start.",
            "filters_header": "Filters",
            "all_years_option": "All Years",
            "select_year": "Select Year",
            "all_months_option": "All Months",
            "select_month": "Select Month",
            "all_projects_option": "All Projects",
            "select_project": "Select Project",
            "overview_report_tab": "Overview Report",
            "comparison_tab": "Comparison",
            "data_preview_tab": "Data Preview",
            "user_guide_tab": "User Guide",
            "overview_report_header": "Project Overview Report",
            "no_data_for_filters": "No data to display with the selected filters.",
            "overall_summary": "Overall Summary",
            "metric_column": "Metric",
            "value_column": "Value",
            "total_hours": "Total Hours Worked",
            "total_cost_usd": "Total Cost (USD)",
            "monthly_summary_header": "Monthly Summary",
            "project_summary_header": "Project Summary",
            "export_report_header": "Export Report",
            "export_excel": "Export to Excel",
            "export_pdf": "Export to PDF",
            "create_report_button": "Create Report",
            "generating_report_spinner": "Generating report...",
            "download_excel_report": "Download Excel Report",
            "download_pdf_report": "Download PDF Report",
            "error_generating_report": "Error generating report. Please check filters and try again.",
            "select_export_format": "Please select at least one export format (Excel or PDF).",
            "comparison_header": "Data Comparison",
            "compare_projects_in_month": "Compare Projects in a Month",
            "compare_projects_in_year": "Compare Projects in a Year",
            "compare_one_project_over_time": "Compare One Project Over Time (Months/Years)",
            "select_comparison_mode": "Select Comparison Mode",
            "comparison_config_header": "Comparison Configuration",
            "comp_month_project_info": "Select ONE year, ONE month, and at least TWO projects.",
            "select_year_single": "Select Year (single)",
            "select_only_one_year": "Please select only ONE year for this mode.",
            "select_month_single": "Select Month (single)",
            "select_only_one_month": "Please select only ONE month for this mode.",
            "select_projects_multiple": "Select Projects (at least 2)",
            "select_at_least_two_projects": "Please select at least TWO projects for this mode.",
            "comp_year_project_info": "Select ONE year and at least TWO projects. Data will be compared by month.",
            "comp_project_over_time_info": "Select ONLY ONE project. Then, select a year (to compare months within that year) OR multiple years (to compare total hours across years).",
            "select_project_single": "Select One Project (single)",
            "select_only_one_project": "Please select ONLY ONE project for this mode.",
            "select_years_or_months_info": "Select 'Year(s)' to compare across months within that year, OR select multiple 'Year(s)' to compare total hours across years.",
            "select_years": "Select Year(s)",
            "select_months_in_year": "Select Month(s) (in selected year)",
            "select_only_one_project_again": "Please select ONLY ONE project.",
            "select_at_least_one_year_or_month": "Please select at least one year OR month.",
            "select_at_least_one_month_if_one_year": "Please select at least one month if you only select one year to compare by month.",
            "cannot_compare_multiple_years_and_months": "Cannot compare multiple years AND months at the same time. Please clear month selection if you want to compare multiple years.",
            "create_comparison_report": "Create Comparison Report",
            "select_at_least_one_project_for_comparison": "Please select at least one project to compare.",
            "generating_comparison_report_spinner": "Generating comparison report...",
            "comparison_chart_header": "Comparison Chart",
            "chart_placeholder": "Chart will be displayed here.", # Placeholder for actual chart
            "comparison_table_header": "Comparison Data Table",
            "download_comparison_excel": "Download Comparison Excel",
            "download_comparison_pdf": "Download Comparison PDF",
            "no_raw_data": "No raw data loaded. Please upload an Excel file.",
            "user_guide": "User Guide",
            "user_guide_content": """
    - Select filters: mode, year, month, project
    - Select report export format (Excel, PDF or both)
    - Click "Create report"
    - Download generated report
    """,
            # New keys for PDF and table headers
            "project_time_report_pdf": "PROJECT TIME REPORT",
            "report_for": "Report For",
            "year": "Year",
            "month": "Month",
            "project": "Project",
            "report_date": "Report Date",
            "month_column": "Month",
            "total_hours_short": "Hours",
            "total_cost_short": "Cost (USD)",
            "project_name_column": "Project Name",
            "comparison_report_title": "PROJECT COMPARISON REPORT",
            "comparison_mode_label": "Comparison Mode",
            "years_label": "Years",
            "months_label": "Months",
            "projects_label": "Projects",
            "hours_by_project": "Hours by Project",
            "hours_by_month_and_project": "Hours by Month and Project",
            "total_hours_over_time": "Total Hours Over Time",
            "chart_not_generated": "Chart could not be generated or is empty.",
            "project_name_col": "Project Name",
            "total_hours_col": "Total Hours",
            "total_cost_col": "Total Cost",
            "filtered_data_sheet": "Filtered Data",
            "monthly_summary_sheet": "Monthly Summary",
            "project_summary_sheet": "Project Summary",
            "total": "Total" # For sum row in comparison table
        },
        "vi": {
            "app_title": "Ứng dụng báo cáo giờ làm việc dự án",
            "upload_excel_file": "Tải lên file Excel (.xlsx)",
            "loading_data_spinner": "Đang tải và xử lý dữ liệu...",
            "file_upload_success": "Tải file thành công!",
            "file_upload_error": "Lỗi khi tải file. Vui lòng kiểm tra định dạng.",
            "upload_file_to_start": "Vui lòng tải lên file Excel để bắt đầu.",
            "filters_header": "Bộ lọc",
            "all_years_option": "Tất cả các năm",
            "select_year": "Chọn năm",
            "all_months_option": "Tất cả các tháng",
            "select_month": "Chọn tháng",
            "all_projects_option": "Tất cả các dự án",
            "select_project": "Chọn dự án",
            "overview_report_tab": "Tổng quan báo cáo",
            "comparison_tab": "So sánh",
            "data_preview_tab": "Xem trước dữ liệu",
            "user_guide_tab": "Hướng dẫn sử dụng",
            "overview_report_header": "Báo cáo tổng quan dự án",
            "no_data_for_filters": "Không có dữ liệu để hiển thị với các bộ lọc đã chọn.",
            "overall_summary": "Tổng quan chung",
            "metric_column": "Chỉ số",
            "value_column": "Giá trị",
            "total_hours": "Tổng giờ làm việc",
            "total_cost_usd": "Tổng chi phí (USD)",
            "monthly_summary_header": "Tổng quan theo tháng",
            "project_summary_header": "Tổng quan theo dự án",
            "export_report_header": "Xuất báo cáo",
            "export_excel": "Xuất ra Excel",
            "export_pdf": "Xuất ra PDF",
            "create_report_button": "Tạo báo cáo",
            "generating_report_spinner": "Đang tạo báo cáo...",
            "download_excel_report": "Tải xuống báo cáo Excel",
            "download_pdf_report": "Tải xuống báo cáo PDF",
            "error_generating_report": "Lỗi khi tạo báo cáo. Vui lòng kiểm tra bộ lọc và thử lại.",
            "select_export_format": "Vui lòng chọn ít nhất một định dạng xuất (Excel hoặc PDF).",
            "comparison_header": "So sánh dữ liệu",
            "compare_projects_in_month": "So sánh dự án trong một tháng",
            "compare_projects_in_year": "So sánh dự án trong một năm",
            "compare_one_project_over_time": "So sánh một dự án qua các tháng/năm",
            "select_comparison_mode": "Chọn chế độ so sánh",
            "comparison_config_header": "Cấu hình so sánh",
            "comp_month_project_info": "Chọn MỘT năm, MỘT tháng và ít nhất HAI dự án.",
            "select_year_single": "Chọn Năm (chỉ 1)",
            "select_only_one_year": "Vui lòng chọn CHỈ MỘT năm cho chế độ này.",
            "select_month_single": "Chọn Tháng (chỉ 1)",
            "select_only_one_month": "Vui lòng chọn CHỈ MỘT tháng cho chế độ này.",
            "select_projects_multiple": "Chọn Các Dự Án (ít nhất 2)",
            "select_at_least_two_projects": "Vui lòng chọn ít nhất HAI dự án cho chế độ này.",
            "comp_year_project_info": "Chọn MỘT năm và ít nhất HAI dự án. Dữ liệu sẽ được so sánh theo tháng.",
            "comp_project_over_time_info": "Chọn CHỈ MỘT dự án. Sau đó, chọn một năm (để so sánh các tháng trong năm đó) HOẶC nhiều năm (để so sánh tổng giờ qua các năm).",
            "select_project_single": "Chọn Một Dự Án (chỉ 1)",
            "select_only_one_project": "Vui lòng chọn CHỈ MỘT dự án cho chế độ này.",
            "select_years_or_months_info": "Chọn 'Năm(s)' để so sánh qua các tháng trong năm đó, HOẶC chọn nhiều 'Năm(s)' để so sánh tổng giờ qua các năm.",
            "select_years": "Chọn Năm(s)",
            "select_months_in_year": "Chọn Tháng(s) (trong năm đã chọn)",
            "select_only_one_project_again": "Vui lòng chọn CHỈ MỘT dự án.",
            "select_at_least_one_year_or_month": "Vui lòng chọn ít nhất một năm HOẶC tháng.",
            "select_at_least_one_month_if_one_year": "Vui lòng chọn ít nhất một tháng nếu bạn chỉ chọn một năm để so sánh theo tháng.",
            "cannot_compare_multiple_years_and_months": "Không thể so sánh nhiều năm VÀ các tháng cùng lúc. Vui lòng xóa lựa chọn tháng nếu bạn muốn so sánh nhiều năm.",
            "create_comparison_report": "Tạo báo cáo so sánh",
            "select_at_least_one_project_for_comparison": "Vui lòng chọn ít nhất một dự án để so sánh.",
            "generating_comparison_report_spinner": "Đang tạo báo cáo so sánh...",
            "comparison_chart_header": "Biểu đồ so sánh",
            "chart_placeholder": "Biểu đồ sẽ hiển thị ở đây.",
            "comparison_table_header": "Bảng dữ liệu so sánh",
            "download_comparison_excel": "Tải xuống Excel so sánh",
            "download_comparison_pdf": "Tải xuống PDF so sánh",
            "no_raw_data": "Chưa tải dữ liệu thô. Vui lòng tải lên file Excel.",
            "user_guide": "Hướng dẫn sử dụng",
            "user_guide_content": """
    - Chọn bộ lọc: chế độ, năm, tháng, dự án
    - Chọn định dạng xuất báo cáo (Excel, PDF hoặc cả hai)
    - Nhấp vào "Tạo báo cáo"
    - Tải xuống báo cáo đã tạo
    """,
            # New keys for PDF and table headers
            "project_time_report_pdf": "BÁO CÁO GIỜ LÀM VIỆC DỰ ÁN",
            "report_for": "Báo cáo cho",
            "year": "Năm",
            "month": "Tháng",
            "project": "Dự án",
            "report_date": "Ngày báo cáo",
            "month_column": "Tháng",
            "total_hours_short": "Giờ",
            "total_cost_short": "Chi phí (USD)",
            "project_name_column": "Tên dự án",
            "comparison_report_title": "BÁO CÁO SO SÁNH DỰ ÁN",
            "comparison_mode_label": "Chế độ so sánh",
            "years_label": "Các năm",
            "months_label": "Các tháng",
            "projects_label": "Các dự án",
            "hours_by_project": "Giờ theo dự án",
            "hours_by_month_and_project": "Giờ theo tháng và dự án",
            "total_hours_over_time": "Tổng giờ theo thời gian",
            "chart_not_generated": "Không thể tạo biểu đồ hoặc không có dữ liệu.",
            "project_name_col": "Tên dự án",
            "total_hours_col": "Tổng giờ",
            "total_cost_col": "Tổng chi phí",
            "filtered_data_sheet": "Dữ liệu đã lọc",
            "monthly_summary_sheet": "Tóm tắt hàng tháng",
            "project_summary_sheet": "Tóm tắt dự án",
            "total": "Tổng" # For sum row in comparison table
        }
    }

def get_text(key, lang_code, lang_data):
    # Fallback logic: 1. specific key in specific lang, 2. key in default lang (en), 3. key itself
    return lang_data.get(lang_code, {}).get(key, lang_data.get("en", {}).get(key, key))

# Hàm hỗ trợ làm sạch tên file/sheet
def sanitize_filename(name):
    # Ký tự không hợp lệ trong tên file/sheet của Excel
    invalid_chars = re.compile(r'[\\\\/*?[\\]:;|=,<>]')
    s = invalid_chars.sub("_", str(name))
    # Loại bỏ các ký tự điều khiển ASCII và các ký tự không an toàn khác
    s = ''.join(c for c in s if c.isprintable())
    return s[:31] # Giới hạn 31 ký tự cho tên sheet trong Excel

def setup_paths():
    """Thiết lập các đường dẫn file đầu vào và đầu ra."""
    today = datetime.datetime.today().strftime('%Y%m%d')
    return {
        'template_file': "Time_report.xlsm",
        'output_file': f"Time_report_Standard_{today}.xlsx",
        'pdf_report': f"Time_report_Standard_{today}.pdf",
        'comparison_output_file': f"Time_report_Comparison_{today}.xlsx",
        'comparison_pdf_report': f"Time_report_Comparison_{today}.pdf",
        'logo_path': "triac_logo.png" # Thêm đường dẫn logo
    }

def read_configs(template_file):
    """Đọc cấu hình từ file template Excel (ví dụ: các giá trị từ các ô cụ thể)."""
    # ... (giữ nguyên logic đọc config của bạn) ...
    # Placeholder for actual implementation based on your Excel template
    return {
        "report_title": "PROJECT TIME REPORT",
        "company_name": "TRIAC GLOBAL",
        "contact_info": "contact@triacglobal.com",
        "version": "1.0.0"
    }

def load_raw_data(uploaded_file, path_dict):
    """Tải và tiền xử lý dữ liệu từ file Excel đã tải lên."""
    try:
        df = pd.read_excel(uploaded_file)
        # Convert date columns to datetime, handling potential errors
        for col in ['Start date', 'End date']:
            df[col] = pd.to_datetime(df[col], errors='coerce')
        
        df.dropna(subset=['Start date', 'End date'], inplace=True) # Drop rows with invalid dates

        df['Month'] = df['Start date'].dt.month
        df['Year'] = df['Start date'].dt.year
        df['MonthName'] = df['Start date'].dt.strftime('%B') # English month names for consistency
        df['Week'] = df['Start date'].dt.isocalendar().week.astype(int)

        # Ensure 'Hours' column is numeric
        df['Hours'] = pd.to_numeric(df['Hours'], errors='coerce')
        df.dropna(subset=['Hours'], inplace=True) # Drop rows where Hours could not be parsed

        return df
    except Exception as e:
        print(f"Error loading or preprocessing data: {e}")
        return pd.DataFrame() # Return empty DataFrame on error

def apply_filters(df_raw, year=None, month_name=None, project_name=None):
    """Áp dụng bộ lọc cho DataFrame."""
    df_filtered = df_raw.copy()
    if year:
        df_filtered = df_filtered[df_filtered['Year'] == year]
    if month_name:
        df_filtered = df_filtered[df_filtered['MonthName'] == month_name]
    if project_name:
        df_filtered = df_filtered[df_filtered['Project name'] == project_name]
    return df_filtered

def export_report(df_filtered, template_file, output_file, year, month_name, project_name, get_text_func):
    """Xuất báo cáo tổng quan ra file Excel (.xlsx)."""
    try:
        if not os.path.exists(template_file):
            print(f"Template file not found: {template_file}")
            # Create a dummy template if not found for basic functionality
            wb = load_workbook()
            ws = wb.active
            ws.title = "Summary"
            wb.save(template_file)
            print(f"Dummy template created at {template_file}")
            
        wb = load_workbook(template_file, keep_vba=True)
        
        # --- Summary Sheet ---
        if 'Summary' in wb.sheetnames:
            ws_summary = wb['Summary']
            # Clear existing data in relevant cells if any, or write to specific cells
            
            # Example: Write filtered data to a new sheet or a predefined range
            # For simplicity, let's write to a new sheet or clear/overwrite existing.
            # You would need to map your summary data to specific cells here.
            # Example:
            # ws_summary['B2'] = df_filtered['Hours'].sum()
            # ws_summary['B3'] = df_filtered['Total cost (USD)'].sum()
            
            # Or insert a new sheet for filtered data
            if get_text_func("filtered_data_sheet") not in wb.sheetnames:
                ws_filtered = wb.create_sheet(get_text_func("filtered_data_sheet")) 
            else:
                ws_filtered = wb[get_text_func("filtered_data_sheet")]
                # Clear existing content if overwriting
                for row in ws_filtered.iter_rows():
                    for cell in row:
                        cell.value = None

            for r_idx, row in enumerate(dataframe_to_rows(df_filtered, index=False, header=True), 1):
                ws_filtered.append(row)

            # Monthly Summary
            monthly_summary = df_filtered.groupby('MonthName').agg(
                total_hours=('Hours', 'sum'),
                total_cost=('Total cost (USD)', 'sum')
            ).reset_index()
            if get_text_func("monthly_summary_sheet") not in wb.sheetnames:
                ws_monthly = wb.create_sheet(get_text_func("monthly_summary_sheet"))
            else:
                ws_monthly = wb[get_text_func("monthly_summary_sheet")]
                for row in ws_monthly.iter_rows():
                    for cell in row:
                        cell.value = None
            for r_idx, row in enumerate(dataframe_to_rows(monthly_summary, index=False, header=True), 1):
                ws_monthly.append(row)

            # Project Summary
            project_summary = df_filtered.groupby('Project name').agg(
                total_hours=('Hours', 'sum'),
                total_cost=('Total cost (USD)', 'sum')
            ).reset_index().sort_values(by='total_hours', ascending=False)
            if get_text_func("project_summary_sheet") not in wb.sheetnames:
                ws_project = wb.create_sheet(get_text_func("project_summary_sheet"))
            else:
                ws_project = wb[get_text_func("project_summary_sheet")]
                for row in ws_project.iter_rows():
                    for cell in row:
                        cell.value = None
            for r_idx, row in enumerate(dataframe_to_rows(project_summary, index=False, header=True), 1):
                ws_project.append(row)


        wb.save(output_file)
        return True
    except Exception as e:
        print(f"Error exporting standard report: {e}")
        return False

# --- PDF Font Setup (using FPDF) ---
# It's better to manage fonts in a dedicated way, e.g., in a 'fonts' folder
# For demonstration, let's assume we place a font like 'DejaVuSansCondensed.ttf'
# in the same directory or a 'fonts' subdirectory.
# You need to download a Unicode font that supports Vietnamese.
# Example: DejaVuSansCondensed.ttf (can be found online, or NotoSans-Regular.ttf from Google Fonts)
# Place this file in a 'fonts' directory next to your script.

FONT_DIR = os.path.join(os.path.dirname(__file__), "fonts")
if not os.path.exists(FONT_DIR):
    os.makedirs(FONT_DIR)
    # You might want to automatically download or prompt user to place font here.

def register_vietnamese_font(pdf):
    font_path = os.path.join(FONT_DIR, "DejaVuSansCondensed.ttf") # Example font
    # Fallback: if the font file doesn't exist, use a generic font.
    # It's highly recommended to ensure 'DejaVuSansCondensed.ttf' (or chosen font) is present.
    if not os.path.exists(font_path):
        print(f"Warning: Font file not found at {font_path}. PDF might not display Vietnamese correctly.")
        pdf.set_font('helvetica', size=12) # Fallback to a standard font
        return False
    try:
        pdf.add_font('DejaVuSans', '', font_path, uni=True)
        # Add bold version if available and needed
        # pdf.add_font('DejaVuSans', 'B', os.path.join(FONT_DIR, "DejaVuSansCondensed-Bold.ttf"), uni=True)
        pdf.set_font('DejaVuSans', size=12)
        print(f"Successfully registered font: {font_path}")
        return True
    except Exception as e:
        print(f"Error registering font {font_path}: {e}. Falling back to Helvetica.")
        pdf.set_font('helvetica', size=12) # Fallback
        return False

def export_pdf_report(df_filtered, pdf_file_path, year, month_name, project_name, logo_path, get_text_func):
    """Xuất báo cáo tổng quan ra PDF."""
    try:
        pdf = FPDF()
        register_vietnamese_font(pdf) # Register font
        
        pdf.add_page()

        # Add logo if exists
        if os.path.exists(logo_path):
            try:
                pdf.image(logo_path, x=10, y=8, w=30)
            except Exception as e:
                print(f"Error adding logo to PDF: {e}")

        # Title
        pdf.ln(10)
        pdf.set_font('DejaVuSans', 'B', 16) # Use registered font for title
        pdf.cell(0, 10, get_text_func("project_time_report_pdf"), ln=True, align='C')
        
        pdf.set_font('DejaVuSans', '', 12) # Use registered font for details
        report_period = []
        if year: report_period.append(f"{get_text_func('year')}: {year}")
        if month_name: report_period.append(f"{get_text_func('month')}: {get_text_func(month_name.lower())}") # Translate month name if available
        if project_name: report_period.append(f"{get_text_func('project')}: {project_name}")
        
        if report_period:
            pdf.cell(0, 7, get_text_func("report_for") + ": " + ", ".join(report_period), ln=True, align='C')
        
        pdf.cell(0, 7, f"{get_text_func('report_date')}: {datetime.datetime.today().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align='C')
        pdf.ln(10)

        # Overall Summary
        pdf.set_font('DejaVuSans', 'B', 14)
        pdf.cell(0, 10, get_text_func("overall_summary"), ln=True)
        pdf.set_font('DejaVuSans', '', 12)

        if df_filtered.empty:
            pdf.cell(0, 10, get_text_func("no_data_for_filters"), ln=True)
        else:
            total_hours = df_filtered['Hours'].sum()
            total_cost = df_filtered['Total cost (USD)'].sum()
            pdf.cell(0, 7, f"{get_text_func('total_hours')}: {total_hours:,.0f}", ln=True)
            pdf.cell(0, 7, f"{get_text_func('total_cost_usd')}: {total_cost:,.2f} USD", ln=True)
        pdf.ln(5)

        # Monthly Summary
        pdf.set_font('DejaVuSans', 'B', 14)
        pdf.cell(0, 10, get_text_func("monthly_summary_header"), ln=True)
        pdf.set_font('DejaVuSans', '', 10)
        if df_filtered.empty:
             pdf.cell(0, 10, get_text_func("no_data_for_filters"), ln=True)
        else:
            monthly_summary = df_filtered.groupby('MonthName').agg(
                total_hours=('Hours', 'sum'),
                total_cost=('Total cost (USD)', 'sum')
            ).reset_index()
            # Order months for display
            month_name_order = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
            ordered_monthly_summary = monthly_summary
            if 'MonthName' in monthly_summary.columns:
                present_months = [m for m in month_name_order if m in monthly_summary['MonthName'].unique()]
                if present_months:
                    ordered_monthly_summary['MonthName'] = pd.Categorical(monthly_summary['MonthName'], categories=present_months, ordered=True)
                    ordered_monthly_summary = ordered_monthly_summary.sort_values('MonthName')

            # Add table headers
            col_widths = [40, 30, 40]
            pdf.set_fill_color(200, 220, 255)
            pdf.cell(col_widths[0], 7, get_text_func("month_column"), 1, 0, 'C', 1)
            pdf.cell(col_widths[1], 7, get_text_func("total_hours_short"), 1, 0, 'C', 1)
            pdf.cell(col_widths[2], 7, get_text_func("total_cost_short"), 1, 1, 'C', 1)
            pdf.set_fill_color(240, 240, 240)

            for index, row in ordered_monthly_summary.iterrows():
                pdf.cell(col_widths[0], 7, get_text_func(row['MonthName'].lower()), 1, 0, 'L') # Translate month name for display
                pdf.cell(col_widths[1], 7, f"{row['total_hours']:,.0f}", 1, 0, 'R')
                pdf.cell(col_widths[2], 7, f"{row['total_cost']:,.2f} USD", 1, 1, 'R')
        pdf.ln(5)

        # Project Summary
        pdf.set_font('DejaVuSans', 'B', 14)
        pdf.cell(0, 10, get_text_func("project_summary_header"), ln=True)
        pdf.set_font('DejaVuSans', '', 10)
        if df_filtered.empty:
            pdf.cell(0, 10, get_text_func("no_data_for_filters"), ln=True)
        else:
            project_summary = df_filtered.groupby('Project name').agg(
                total_hours=('Hours', 'sum'),
                total_cost=('Total cost (USD)', 'sum')
            ).reset_index().sort_values(by='total_hours', ascending=False)
            
            col_widths_proj = [80, 30, 40]
            pdf.set_fill_color(200, 220, 255)
            pdf.cell(col_widths_proj[0], 7, get_text_func("project_name_column"), 1, 0, 'C', 1)
            pdf.cell(col_widths_proj[1], 7, get_text_func("total_hours_short"), 1, 0, 'C', 1)
            pdf.cell(col_widths_proj[2], 7, get_text_func("total_cost_short"), 1, 1, 'C', 1)
            pdf.set_fill_color(240, 240, 240)

            for index, row in project_summary.iterrows():
                pdf.cell(col_widths_proj[0], 7, row['Project name'], 1, 0, 'L')
                pdf.cell(col_widths_proj[1], 7, f"{row['total_hours']:,.0f}", 1, 0, 'R')
                pdf.cell(col_widths_proj[2], 7, f"{row['total_cost']:,.2f} USD", 1, 1, 'R')
        
        pdf.output(pdf_file_path, 'F')
        print(f"Standard PDF report saved to {pdf_file_path}")
        return True
    except Exception as e:
        print(f"Error exporting PDF report: {e}")
        return False

# --- Comparison Functions ---
def apply_comparison_filters(df_raw, comparison_config, comparison_mode, get_text_func):
    """Áp dụng bộ lọc và tạo DataFrame tóm tắt cho báo cáo so sánh."""
    years = comparison_config.get('years', [])
    months = comparison_config.get('months', [])
    selected_projects = comparison_config.get('selected_projects', [])

    df_filtered_comp = df_raw.copy()

    if years:
        df_filtered_comp = df_filtered_comp[df_filtered_comp['Year'].isin(years)]
    
    if months:
        df_filtered_comp = df_filtered_comp[df_filtered_comp['MonthName'].isin(months)]
    
    if selected_projects:
        df_filtered_comp = df_filtered_comp[df_filtered_comp['Project name'].isin(selected_projects)]
    else: 
        return pd.DataFrame(), get_text_func("select_at_least_one_project_for_comparison"), pd.DataFrame(), pd.DataFrame()

    if df_filtered_comp.empty:
        return pd.DataFrame(), get_text_func("no_data_for_filters"), pd.DataFrame(), pd.DataFrame()

    chart_data = pd.DataFrame() # Data specifically for chart
    table_data = pd.DataFrame() # Data specifically for table display
    msg = ""

    month_name_order = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]

    if comparison_mode == "month_project":
        if len(years) != 1 or len(months) != 1 or len(selected_projects) < 2:
            return pd.DataFrame(), get_text_func("comp_month_project_info"), pd.DataFrame(), pd.DataFrame()
        
        chart_data = df_filtered_comp.groupby('Project name')['Hours'].sum().reset_index()
        chart_data.rename(columns={'Hours': get_text_func("total_hours_col")}, inplace=True)
        table_data = chart_data.copy()
        msg = f"{get_text_func('compare_projects_in_month')} ({months[0]}, {years[0]})"
        return chart_data, msg, chart_data, table_data

    elif comparison_mode == "year_project":
        if len(years) != 1 or len(selected_projects) < 2:
            return pd.DataFrame(), get_text_func("comp_year_project_info"), pd.DataFrame(), pd.DataFrame()
        
        table_data = df_filtered_comp.groupby(['Project name', 'MonthName'])['Hours'].sum().unstack(fill_value=0)
        
        existing_months = [m for m in month_name_order if m in table_data.columns]
        table_data = table_data[existing_months]

        table_data = table_data.reset_index().rename(columns={'index': get_text_func("project_name_col")})
        
        table_data[get_text_func("total_hours_col")] = table_data[existing_months].sum(axis=1)

        if not table_data.empty:
            table_data.loc[get_text_func("total")] = table_data[existing_months + [get_text_func("total_hours_col")]].sum()
            table_data.loc[get_text_func("total"), get_text_func("project_name_col")] = get_text_func("total")
        
        # Prepare chart_data (melted for easy plotting)
        chart_data = table_data[table_data[get_text_func("project_name_col")] != get_text_func("total")].melt(
            id_vars=[get_text_func("project_name_col")], value_vars=existing_months, 
            var_name='MonthName', value_name='Hours'
        )
        if 'MonthName' in chart_data.columns:
            chart_data['MonthName'] = pd.Categorical(chart_data['MonthName'], categories=existing_months, ordered=True)
            chart_data = chart_data.sort_values('MonthName')

        msg = f"{get_text_func('compare_projects_in_year')} ({years[0]} {get_text_func('by_month')})"
        return table_data, msg, chart_data, table_data # Return table_data as df_comparison and actual table_data

    elif comparison_mode == "project_over_time":
        if len(selected_projects) != 1:
            return pd.DataFrame(), get_text_func("select_only_one_project"), pd.DataFrame(), pd.DataFrame()

        selected_project_name = selected_projects[0]

        if len(years) == 1 and len(months) > 0:
            chart_data = df_filtered_comp.groupby('MonthName')['Hours'].sum().reset_index()
            chart_data.rename(columns={'Hours': f'{get_text_func("total_hours_col")} {get_text_func("for_project")} {selected_project_name}'}, inplace=True)
            
            if not chart_data.empty and 'MonthName' in chart_data.columns:
                present_months = [m for m in month_name_order if m in chart_data['MonthName'].unique()]
                if present_months: 
                    chart_data['MonthName'] = pd.Categorical(chart_data['MonthName'], categories=present_months, ordered=True)
                    chart_data = chart_data.sort_values('MonthName').reset_index(drop=True)
            else:
                return pd.DataFrame(), f"{get_text_func('no_month_data_for_project')} '{selected_project_name}' {get_text_func('in_year')} {years[0]}.", pd.DataFrame(), pd.DataFrame()

            chart_data[get_text_func('project_name_col')] = selected_project_name
            table_data = chart_data.copy()
            msg = f"{get_text_func('total_hours_project')} {selected_project_name} {get_text_func('over_months_in_year')} {years[0]}"
            return chart_data, msg, chart_data, table_data

        elif len(years) > 1 and not months:
            chart_data = df_filtered_comp.groupby('Year')['Hours'].sum().reset_index()
            chart_data.rename(columns={'Hours': f'{get_text_func("total_hours_col")} {get_text_func("for_project")} {selected_project_name}'}, inplace=True)
            chart_data['Year'] = chart_data['Year'].astype(str)
            
            chart_data[get_text_func('project_name_col')] = selected_project_name
            table_data = chart_data.copy()
            msg = f"{get_text_func('total_hours_project')} {selected_project_name} {get_text_func('over_years')}"
            return chart_data, msg, chart_data, table_data

        else:
            return pd.DataFrame(), get_text_func("invalid_comparison_config_time"), pd.DataFrame(), pd.DataFrame()
        
    return pd.DataFrame(), get_text_func("invalid_comparison_mode"), pd.DataFrame(), pd.DataFrame()


def export_comparison_report(df_comparison, comparison_mode, output_file, comparison_config, get_text_func):
    """Xuất báo cáo so sánh ra file Excel (.xlsx)."""
    try:
        wb = load_workbook() # Create a new workbook
        ws = wb.active
        ws.title = get_text_func("comparison_report_sheet") # Use translated sheet name

        # Add header for the report
        ws.append([get_text_func("comparison_report_title")])
        ws.append([f"{get_text_func('comparison_mode_label')}: {get_text_func(comparison_mode)}"])
        if comparison_config.get('years'):
            ws.append([f"{get_text_func('years_label')}: {', '.join(map(str, comparison_config['years']))}"])
        if comparison_config.get('months'):
            ws.append([f"{get_text_func('months_label')}: {', '.join(comparison_config['months'])}"])
        if comparison_config.get('selected_projects'):
            ws.append([f"{get_text_func('projects_label')}: {', '.join(comparison_config['selected_projects'])}"])
        ws.append([]) # Empty row for spacing

        # Write DataFrame to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df_comparison, index=False, header=True), len(ws['A']) + 1):
            ws.append(row)
        
        # Apply some basic formatting (optional)
        for cell in ws[1]: # First row (title)
            cell.font = cell.font.copy(bold=True)
        
        wb.save(output_file)
        return True
    except Exception as e:
        print(f"Error exporting comparison Excel report: {e}")
        return False

def create_chart_image(df, chart_type, x_col, y_col, color_col=None, title="", output_file="chart.png"):
    """
    Creates a chart using matplotlib and saves it as a PNG image.
    This function is primarily for generating images for PDF export.
    Returns path to the saved image or None if failed.
    """
    if df.empty:
        print(f"Warning: Empty DataFrame provided for chart generation: {title}")
        # Create a blank image or an image with "No Data" text
        plt.figure(figsize=(10, 6))
        plt.text(0.5, 0.5, "No Data Available", horizontalalignment='center', verticalalignment='center', fontsize=20, color='gray', transform=plt.gca().transAxes)
        plt.axis('off') # Hide axes
        plt.title(title)
        try:
            plt.savefig(output_file, format='png')
            plt.close()
            return output_file
        except Exception as e:
            print(f"Error saving empty chart placeholder: {e}")
            plt.close()
            return None

    try:
        plt.figure(figsize=(10, 6))
        
        if chart_type == 'bar':
            if color_col and color_col in df.columns:
                # Grouped bar chart
                df.pivot_table(index=x_col, columns=color_col, values=y_col).plot(kind='bar', ax=plt.gca())
            else:
                plt.bar(df[x_col], df[y_col])
        elif chart_type == 'line':
            if color_col and color_col in df.columns:
                for name, group in df.groupby(color_col):
                    plt.plot(group[x_col], group[y_col], label=name, marker='o')
                plt.legend()
            else:
                plt.plot(df[x_col], df[y_col], marker='o')
        
        plt.title(title)
        plt.xlabel(x_col)
        plt.ylabel(y_col)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        plt.savefig(output_file, format='png')
        plt.close()
        return output_file
    except Exception as e:
        print(f"Error creating chart image: {e}")
        # Clean up if partially created
        if os.path.exists(output_file):
            os.remove(output_file)
        plt.close() # Ensure plot is closed even on error
        return None

def export_comparison_pdf_report(chart_df, table_df, comparison_mode, pdf_file_path, logo_path, comparison_config, get_text_func):
    """Xuất báo cáo so sánh ra PDF."""
    temp_chart_path = os.path.join(tempfile.gettempdir(), "comparison_chart.png")
    
    try:
        pdf = FPDF()
        register_vietnamese_font(pdf) # Register font
        
        pdf.add_page()

        # Add logo if exists
        if os.path.exists(logo_path):
            try:
                pdf.image(logo_path, x=10, y=8, w=30)
            except Exception as e:
                print(f"Error adding logo to PDF: {e}")

        pdf.ln(10)
        pdf.set_font('DejaVuSans', 'B', 16)
        pdf.cell(0, 10, get_text_func("comparison_report_title"), ln=True, align='C')
        
        pdf.set_font('DejaVuSans', '', 12)
        mode_text = ""
        if comparison_mode == "month_project":
            mode_text = get_text_func("compare_projects_in_month")
        elif comparison_mode == "year_project":
            mode_text = get_text_func("compare_projects_in_year")
        elif comparison_mode == "project_over_time":
            mode_text = get_text_func("compare_one_project_over_time")
        
        pdf.cell(0, 7, f"{get_text_func('comparison_mode_label')}: {mode_text}", ln=True, align='C')
        
        config_info_lines = []
        if comparison_config.get('years'):
            config_info_lines.append(f"{get_text_func('years_label')}: {', '.join(map(str, comparison_config['years']))}")
        if comparison_config.get('months'):
            translated_months = [get_text_func(m.lower()) for m in comparison_config['months']] # Translate month names
            config_info_lines.append(f"{get_text_func('months_label')}: {', '.join(translated_months)}")
        if comparison_config.get('selected_projects'):
            config_info_lines.append(f"{get_text_func('projects_label')}: {', '.join(comparison_config['selected_projects'])}")
        
        for line in config_info_lines:
            pdf.cell(0, 7, line, ln=True, align='C')

        pdf.cell(0, 7, f"{get_text_func('report_date')}: {datetime.datetime.today().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align='C')
        pdf.ln(10)

        # Add Chart Section (as image)
        chart_title_val = ""
        chart_x_col = ""
        chart_y_col = ""
        chart_color_col = None
        chart_type = 'bar'
        
        chart_data_for_plot = chart_df.copy() # Use the data returned specifically for the chart

        if comparison_mode == "month_project":
            chart_title_val = get_text_func("comparison_chart_header") + " - " + get_text_func("hours_by_project")
            chart_x_col = 'Project name'
            chart_y_col = get_text_func("total_hours_col")
            created_chart_path = create_chart_image(chart_data_for_plot, 'bar', chart_x_col, chart_y_col, title=chart_title_val, output_file=temp_chart_path)
        elif comparison_mode == "year_project":
            chart_title_val = get_text_func("comparison_chart_header") + " - " + get_text_func("hours_by_month_and_project")
            chart_x_col = 'MonthName'
            chart_y_col = 'Hours'
            chart_color_col = get_text_func("project_name_col")
            # Ensure MonthName is categorical for correct order in chart_data_for_plot
            month_name_order = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
            if 'MonthName' in chart_data_for_plot.columns:
                present_months = [m for m in month_name_order if m in chart_data_for_plot['MonthName'].unique()]
                if present_months:
                    chart_data_for_plot['MonthName'] = pd.Categorical(chart_data_for_plot['MonthName'], categories=present_months, ordered=True)
                    chart_data_for_plot = chart_data_for_plot.sort_values('MonthName')

            created_chart_path = create_chart_image(chart_data_for_plot, 'bar', chart_x_col, chart_y_col, color_col=chart_color_col, title=chart_title_val, output_file=temp_chart_path)
        elif comparison_mode == "project_over_time":
            chart_title_val = get_text_func("comparison_chart_header") + " - " + get_text_func("total_hours_over_time")
            
            if 'MonthName' in chart_data_for_plot.columns:
                chart_x_col = 'MonthName'
                # Ensure MonthName is categorical for correct order
                month_name_order = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
                if 'MonthName' in chart_data_for_plot.columns:
                    present_months = [m for m in month_name_order if m in chart_data_for_plot['MonthName'].unique()]
                    if present_months:
                        chart_data_for_plot['MonthName'] = pd.Categorical(chart_data_for_plot['MonthName'], categories=present_months, ordered=True)
                        chart_data_for_plot = chart_data_for_plot.sort_values('MonthName')
            elif 'Year' in chart_data_for_plot.columns:
                chart_x_col = 'Year'
            chart_y_col = chart_data_for_plot.columns[1] # Assuming the hours column is the second one
            created_chart_path = create_chart_image(chart_data_for_plot, 'line', chart_x_col, chart_y_col, title=chart_title_val, output_file=temp_chart_path)
        else:
            created_chart_path = None # No chart for unknown mode

        pdf.set_font('DejaVuSans', 'B', 14)
        pdf.cell(0, 10, get_text_func("comparison_chart_header"), ln=True)
        if created_chart_path and os.path.exists(created_chart_path):
            try:
                pdf.image(created_chart_path, x=10, w=180) # Adjust width as needed
            except Exception as e:
                print(f"Error adding chart image to PDF: {e}")
                pdf.set_font('DejaVuSans', '', 12)
                pdf.cell(0, 10, get_text_func("chart_not_generated"), ln=True)
        else:
            pdf.set_font('DejaVuSans', '', 12)
            pdf.cell(0, 10, get_text_func("chart_not_generated"), ln=True)
        pdf.ln(5)

        # Add Table Section
        pdf.set_font('DejaVuSans', 'B', 14)
        pdf.cell(0, 10, get_text_func("comparison_table_header"), ln=True)
        pdf.set_font('DejaVuSans', '', 10)

        if table_df.empty:
            pdf.cell(0, 10, get_text_func("no_data_for_filters"), ln=True)
        else:
            # Prepare table data for FPDF
            headers = table_df.columns.tolist()
            data_rows = table_df.values.tolist()

            # Calculate column widths dynamically or set fixed
            col_widths = []
            for col in headers:
                # Use translated header if available for width calculation, but fall back to original
                translated_col = get_text_func(col.replace(" ", "_").lower()) 
                header_text_for_width = translated_col if translated_col != col.replace(" ", "_").lower() else col
                
                max_len_in_col = max(table_df[col].astype(str).apply(len).max(), len(header_text_for_width))
                col_widths.append(max(20, min(50, max_len_in_col * 2))) # Min 20, Max 50, scale by char length

            total_width = sum(col_widths)
            page_width = pdf.w - 2*pdf.l_margin # Approx. usable width
            if total_width > page_width:
                scale_factor = page_width / total_width
                col_widths = [w * scale_factor for w in col_widths]

            pdf.set_fill_color(200, 220, 255)
            # Add table headers
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 7, get_text_func(header.replace(" ", "_").lower()) if get_text_func(header.replace(" ", "_").lower()) != header.replace(" ", "_").lower() else header, 1, 0, 'C', 1)
            pdf.ln()

            pdf.set_fill_color(240, 240, 240)
            # Add table rows
            for row in data_rows:
                for i, item in enumerate(row):
                    # Format numbers
                    if isinstance(item, (int, float)):
                        if headers[i] == get_text_func("total_hours_col") or headers[i] == "Hours": # Check both original and translated
                             item_str = f"{item:,.0f}"
                        elif "Cost" in headers[i] or headers[i] == get_text_func("total_cost_col"):
                             item_str = f"{item:,.2f} USD"
                        else:
                             item_str = str(item)
                    else:
                        item_str = str(item)
                    
                    # Translate month names in row if applicable (for 'year_project' comparison mode)
                    if headers[i] == 'MonthName' and item_str in ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]:
                        item_str = get_text_func(item_str.lower())
                    
                    pdf.cell(col_widths[i], 7, item_str, 1, 0, 'L')
                pdf.ln()

        pdf.output(pdf_file_path, 'F')
        print(f"Comparison PDF report saved to {pdf_file_path}")
        return True
    except Exception as e:
        print(f"Error exporting comparison PDF report: {e}")
        return False
    finally:
        # Clean up temporary chart file
        if os.path.exists(temp_chart_path):
            os.remove(temp_chart_path)
            print(f"Cleaned up temporary chart file: {temp_chart_path}")
