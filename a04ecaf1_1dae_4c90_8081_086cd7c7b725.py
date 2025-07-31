import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils.dataframe import dataframe_to_rows
from fpdf import FPDF
from matplotlib import pyplot as plt
import tempfile
import re
import shutil
from pandas import Series
import traceback
import numpy as np
import smtplib
from email.mime.text import MIMEText

# Hàm hỗ trợ làm sạch tên file/sheet
def sanitize_filename(name):
    # Ký tự không hợp lệ trong tên file/sheet của Excel
    invalid_chars = re.compile(r'[\\/*?[\]:;|=,<>]')
    s = invalid_chars.sub("_", str(name))
    # Loại bỏ các ký tự điều khiển ASCII và các ký tự không an toàn khác
    s = ''.join(c for c in s if c.isprintable())
    return s[:31] # Giới hạn 31 ký tự cho tên sheet trong Excel

def setup_paths():
    """Thiết lập các đường dẫn file đầu vào và đầu ra."""
    today = datetime.today().strftime('%Y%m%d')
    return {
        'template_file': "Time_report.xlsm",
        'output_file': f"Time_report_Standard_{today}.xlsx",
        'pdf_report': f"Time_report_Standard_{today}.pdf",
        'comparison_output_file': f"Time_report_Comparison_{today}.xlsx",
        'comparison_pdf_report': f"Time_report_Comparison_{today}.pdf",
        'logo_path': "triac_logo.png" # Thêm đường dẫn logo
    }
def get_comparison_pdf_path(comparison_mode, base_path):
    if comparison_mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
        return base_path.replace(".pdf", "_Month.pdf")
    elif comparison_mode in ["So Sánh Một Dự Án Qua Các Tháng/Năm", "Compare One Project Over Time (Months/Years)"]:
        return base_path.replace(".pdf", "_SingleProjMonths.pdf")
    elif comparison_mode in ["So Sánh Một Dự Án Qua Các Năm", "Compare One Project Over Years"]:
        return base_path.replace(".pdf", "_SingleProjYears.pdf")
    else:
        return base_path
        
def get_comparison_excel_path(comparison_mode, base_path):
    if comparison_mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
        return base_path.replace(".xlsx", "_Month.xlsx")
    elif comparison_mode in ["So Sánh Một Dự Án Qua Các Tháng/Năm", "Compare One Project Over Time (Months/Years)"]:
        return base_path.replace(".xlsx", "_SingleProjMonths.xlsx")
    elif comparison_mode in ["So Sánh Một Dự Án Qua Các Năm", "Compare One Project Over Years"]:
        return base_path.replace(".xlsx", "_SingleProjYears.xlsx")
    else:
        return base_path

def read_configs(template_file):
    """Đọc cấu hình từ file template Excel."""
    try:
        year_mode_df = pd.read_excel(template_file, sheet_name='Config_Year_Mode', engine='openpyxl')
        project_filter_df = pd.read_excel(template_file, sheet_name='Config_Project_Filter', engine='openpyxl')

        # Xử lý mode, year, months an toàn hơn
        mode_row = year_mode_df.loc[year_mode_df['Key'].str.lower() == 'mode', 'Value']
        mode = str(mode_row.values[0]).strip().lower() if not mode_row.empty and pd.notna(mode_row.values[0]) else 'year'

        year_row = year_mode_df.loc[year_mode_df['Key'].str.lower() == 'year', 'Value']
        year = int(year_row.values[0]) if not year_row.empty and pd.notna(year_row.values[0]) and pd.api.types.is_number(year_row.values[0]) else datetime.datetime.now().year
        
        months_row = year_mode_df.loc[year_mode_df['Key'].str.lower() == 'months', 'Value']
        months = [m.strip().capitalize() for m in str(months_row.values[0]).split(',')] if not months_row.empty and pd.notna(months_row.values[0]) else []
        
        if 'Include' in project_filter_df.columns:
            project_filter_df['Include'] = project_filter_df['Include'].astype(str).str.lower()

        return {
            'mode': mode,
            'year': year,
            'months': months,
            'project_filter_df': project_filter_df
        }
    except FileNotFoundError:
        print(f"Lỗi: Không tìm thấy file template tại {template_file}")
        return {'mode': 'year', 'year': datetime.datetime.now().year, 'months': [], 'project_filter_df': pd.DataFrame(columns=['Project Name', 'Include'])}
    except Exception as e:
        print(f"Lỗi khi đọc cấu hình: {e}")
        return {'mode': 'year', 'year': datetime.datetime.now().year, 'months': [], 'project_filter_df': pd.DataFrame(columns=['Project Name', 'Include'])}

def load_raw_data(template_file):
    """Tải dữ liệu thô từ file template Excel."""
    try:
        df = pd.read_excel(template_file, sheet_name='Raw Data', engine='openpyxl')
        df.columns = df.columns.str.strip()
        df.rename(columns={'Hou': 'Hours', 'Team member': 'Employee', 'Project Name': 'Project name'}, inplace=True)
        
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.dropna(subset=['Date']) # Loại bỏ hàng không có ngày hợp lệ
        
        df['Year'] = df['Date'].dt.year
        df['MonthName'] = df['Date'].dt.month_name()
        df['Week'] = df['Date'].dt.isocalendar().week.astype(int)
        
        # Đảm bảo cột 'Hours' là số
        df['Hours'] = pd.to_numeric(df['Hours'], errors='coerce').fillna(0)
        
        return df
    except Exception as e:
        print(f"Lỗi khi tải dữ liệu thô: {e}")
        return pd.DataFrame()

def apply_filters(df, config):
    """Áp dụng các bộ lọc dữ liệu dựa trên cấu hình."""
    df_filtered = df.copy()

    # ✅ Lọc theo nhiều năm nếu có
    if 'years' in config and config['years']:  # Dành cho so sánh nhiều năm
        df_filtered = df_filtered[df_filtered['Year'].isin(config['years'])]

    # ✅ Lọc theo 1 hoặc nhiều năm nếu là báo cáo tiêu chuẩn
    elif 'year' in config and config['year']:  # Dành cho báo cáo tiêu chuẩn
        if isinstance(config['year'], list):
            df_filtered = df_filtered[df_filtered['Year'].isin(config['year'])]
        else:
            df_filtered = df_filtered[df_filtered['Year'] == config['year']]

    # ✅ Lọc theo tháng (nếu có)
    if config.get('months'):
        df_filtered = df_filtered[df_filtered['MonthName'].isin(config['months'])]

    # ✅ Lọc theo project
    if not config['project_filter_df'].empty:
        selected_project_names = config['project_filter_df']['Project Name'].tolist()
        df_filtered = df_filtered[df_filtered['Project name'].isin(selected_project_names)]
    else:
        return pd.DataFrame(columns=df.columns)  # Trả dataframe rỗng nếu không có project

    return df_filtered

def export_report(df, config, output_file_path):
    """Xuất báo cáo tiêu chuẩn ra file Excel."""
    mode = config.get('mode', 'year')
    
    groupby_cols = []
    if mode == 'year':
        groupby_cols = ['Year', 'Project name']
    elif mode == 'month':
        groupby_cols = ['Year', 'MonthName', 'Project name']
    else: # week mode
        groupby_cols = ['Year', 'Week', 'Project name']

    for col in groupby_cols + ['Hours']:
        if col not in df.columns:
            print(f"Lỗi: Cột '{col}' không tồn tại trong DataFrame. Không thể tạo báo cáo.")
            return False

    if df.empty:
        print("Cảnh báo: DataFrame đã lọc trống, không có báo cáo nào được tạo.")
        return False

    summary = df.groupby(groupby_cols)['Hours'].sum().reset_index()

    try:
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            # Ghi summary cơ bản để giữ nguyên dòng xử lý
            df.to_excel(writer, sheet_name='RawData', index=False)

        wb = load_workbook(output_file_path)

        # === Ghi summary dạng MonthName - Hours ===
        summary_chart = df.groupby('MonthName')['Hours'].sum().reset_index()
        summary_chart = summary_chart.sort_values('MonthName', key=lambda x: pd.to_datetime(x, format='%B'))

        if 'Summary' in wb.sheetnames:
            ws = wb['Summary']
            wb.remove(ws)
        ws = wb.create_sheet("Summary", 0)

        ws.append(['MonthName', 'Hours'])
        for row in summary_chart.itertuples(index=False):
            ws.append([row.MonthName, row.Hours])

        # Thêm biểu đồ vào sheet Summary
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=1 + len(summary_chart))
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=1 + len(summary_chart))

        chart = BarChart()
        chart.title = "Total Hours by Month"
        chart.x_axis.title = "Month"
        chart.y_axis.title = "Hours"
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "E2")

        for project in df['Project name'].unique():
            df_proj = df[df['Project name'] == project]
            sheet_title = sanitize_filename(project)
            
            if sheet_title in wb.sheetnames:
                ws_proj = wb[sheet_title]
            else:
                ws_proj = wb.create_sheet(title=sheet_title)

            summary_task = df_proj.groupby('Task')['Hours'].sum().reset_index().sort_values('Hours', ascending=False)
            
            if not summary_task.empty:
                ws_proj.append(['Task', 'Hours'])
                for row_data in dataframe_to_rows(summary_task, index=False, header=False):
                    ws_proj.append(row_data)

                chart_task = BarChart()
                chart_task.title = f"{project} - Hours by Task"
                chart_task.x_axis.title = "Task"
                chart_task.y_axis.title = "Hours"
                task_len = len(summary_task)
                
                data_ref_task = Reference(ws_proj, min_col=2, min_row=1, max_row=task_len + 1)
                cats_ref_task = Reference(ws_proj, min_col=1, min_row=2, max_row=task_len + 1)
                chart_task.add_data(data_ref_task, titles_from_data=True)
                chart_task.set_categories(cats_ref_task)
                ws_proj.add_chart(chart_task, f"E1")

            start_row_raw_data = ws_proj.max_row + 2 if ws_proj.max_row > 1 else 1
            if not summary_task.empty:
                start_row_raw_data += 15

            for r_idx, r in enumerate(dataframe_to_rows(df_proj, index=False, header=True)):
                for c_idx, cell_val in enumerate(r):
                    ws_proj.cell(row=start_row_raw_data + r_idx, column=c_idx + 1, value=cell_val)
        
        ws_config = wb.create_sheet("Config_Info")
        ws_config['A1'], ws_config['B1'] = "Mode", config.get('mode', 'N/A').capitalize()
        ws_config['A2'], ws_config['B2'] = "Year(s)", ', '.join(map(str, config.get('years', []))) if config.get('years') else str(config.get('year', 'N/A'))
        ws_config['A3'], ws_config['B3'] = "Months", ', '.join(config.get('months', [])) if config.get('months') else "All"
        
        if 'project_filter_df' in config and not config['project_filter_df'].empty:
            selected_projects_display = config['project_filter_df'][config['project_filter_df']['Include'].astype(str).str.lower() == 'yes']['Project Name'].tolist()
            ws_config['A4'], ws_config['B4'] = "Projects Included", ', '.join(selected_projects_display)
        else:
            ws_config['A4'], ws_config['B4'] = "Projects Included", "No projects selected or found"

        # Remove template sheets
        for sheet_name in ['Raw Data', 'Config_Year_Mode', 'Config_Project_Filter']:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

        wb.save(output_file_path)
        return True
    except Exception as e:
        print(f"Lỗi khi xuất báo cáo tiêu chuẩn: {e}")
        return False


def export_pdf_report(df, config, pdf_report_path, logo_path):
    """Xuất báo cáo PDF tiêu chuẩn với các biểu đồ."""
    if not pdf_report_path:
        raise ValueError("❌ pdf_report_path is empty. Please check where it's defined.")
        
    tmp_dir = tempfile.mkdtemp()
    charts_for_pdf = []

    try:
        print(f"[DEBUG] Đường dẫn PDF sẽ ghi: {pdf_report_path}")
        # Tạo biểu đồ tổng số giờ theo tháng
        print(f"[DEBUG] Cột của df: {df.columns.tolist()}")
        if 'MonthName' not in df.columns or 'Hours' not in df.columns:
            raise ValueError("⚠️ Thiếu cột 'MonthName' hoặc 'Hours' trong dữ liệu. Không thể tạo biểu đồ.")
            
        summary_chart = df.groupby('MonthName')['Hours'].sum().reset_index()
        summary_chart = summary_chart.sort_values('MonthName', key=lambda x: pd.to_datetime(x, format='%B'))

        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.bar(summary_chart['MonthName'], summary_chart['Hours'], color='skyblue')  # <- gán vào biến bars
        ax.set_title("Tổng giờ theo tháng")
        ax.set_xlabel("Tháng")
        ax.set_ylabel("Giờ")
        # ✅ Thêm nhãn số giờ trên đầu mỗi cột
        ax.bar_label(bars, labels=[f"{v:.1f}" for v in summary_chart['Hours']], padding=3)
        
        plt.xticks(rotation=45)
        plt.tight_layout()
        chart_path = os.path.join(tmp_dir, "standard_month_chart.png")
        fig.savefig(chart_path, dpi=150)
        plt.close(fig)

        charts_for_pdf.append((chart_path, "Total hour by month", None))
        # 🟩 Thêm biểu đồ Workcentre & Task theo từng dự án
        if 'Project name' in df.columns:
            for project in df['Project name'].dropna().unique():
                safe_project = sanitize_filename(project)
                df_proj = df[df['Project name'] == project]
            # Workcentre
                if 'Workcentre' in df_proj.columns and not df_proj['Workcentre'].empty:
                    wc_summary = df_proj.groupby('Workcentre')['Hours'].sum().sort_values(ascending=False)
                    if not wc_summary.empty and wc_summary.sum() > 0:
                        fig, ax = plt.subplots(figsize=(10, 5))
                        bars = ax.barh(wc_summary.index, wc_summary.values, color='skyblue')
                        ax.bar_label(bars, labels=[f"{v:.1f}" for v in wc_summary.values], padding=3)
                        ax.set_title(f"{project} - Hours by Workcentre", fontsize=10)
                        ax.tick_params(axis='y', labelsize=8)
                        ax.set_xlabel("Hours")
                        ax.set_ylabel("Workcentre")
                        wc_path = os.path.join(tmp_dir, f"{safe_project}_wc.png")
                        plt.tight_layout()
                        fig.savefig(wc_path, dpi=150)
                        plt.close(fig)
                        charts_for_pdf.append((wc_path, f"{project} - Hours by Workcentre", project))
                # Task
                if 'Task' in df_proj.columns and not df_proj['Task'].empty:
                    task_summary = df_proj.groupby('Task')['Hours'].sum().sort_values(ascending=False)
                    if not task_summary.empty and task_summary.sum() > 0:
                        fig, ax = plt.subplots(figsize=(10, 6))
                        bars = ax.barh(task_summary.index, task_summary.values, color='lightgreen')
                        ax.bar_label(bars, labels=[f"{v:.1f}" for v in task_summary.values], padding=3)
                        ax.set_title(f"{project} - Hours by Task", fontsize=10)
                        ax.tick_params(axis='y', labelsize=8)
                        ax.set_xlabel("Hours")
                        ax.set_ylabel("Task")
                        task_path = os.path.join(tmp_dir, f"{safe_project}_task.png")
                        plt.tight_layout()
                        fig.savefig(task_path, dpi=150)
                        plt.close(fig)
                        charts_for_pdf.append((task_path, f"{project} - Hours by Task", project))
        pdf_config_info = {
            "Mode": config.get('mode', 'N/A').capitalize(),
            "Year": str(config.get('year', '')),
            "Months": ', '.join(config.get('months', [])) if config.get('months') else "Tất cả",
            "Project": ', '.join(
                config['project_filter_df'][
                    config['project_filter_df']['Include'] == 'yes'
                ]['Project Name'].tolist()
            ) if 'project_filter_df' in config and not config['project_filter_df'].empty else "Không có"
        }

        success, msg = create_pdf_from_charts_comp(
            charts_for_pdf,
            pdf_report_path,
            "TRIAC TIME REPORT - STANDARD",
            pdf_config_info,
            logo_path
        )
        print(f"[DEBUG] PDF export success: {success}, message: {msg}")
        return success
    except Exception as e:
        print(f"❌ Lỗi khi tạo báo cáo PDF tiêu chuẩn: {e}")
        traceback.print_exc()  # in ra full stacktrace
        return False
    finally:
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir)


def create_pdf_from_charts_comp(charts_data, output_path, title, config_info, logo_path_inner, filter_mode="Total"):
    from collections import defaultdict
    from PIL import Image
    today_str = datetime.today().strftime('%Y-%m-%d')
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    # ✅ Font
    pdf.add_font('DejaVu', '', 'font/dejavu-fonts-ttf-2.37/ttf/DejaVuSans.ttf', uni=True)
    pdf.add_font('DejaVu', 'B', 'font/dejavu-fonts-ttf-2.37/ttf/DejaVuSans-Bold.ttf', uni=True)

    # =========================
    # 🟨 Trang đầu: tiêu đề + thông tin
    # =========================
    pdf.set_font('DejaVu', 'B', 16)
    pdf.add_page()
    if os.path.exists(logo_path_inner):
        pdf.image(logo_path_inner, x=10, y=10, w=30)
    pdf.ln(35)
    pdf.cell(0, 10, title, ln=True, align='C')

    pdf.set_font("DejaVu", '', 11)
    pdf.ln(5)
    pdf.cell(0, 10, f"Generated on: {today_str}", ln=True, align='C')
    pdf.ln(10)

    # Bảng thông tin
    label_width = 40
    value_width = 150
    line_height = 8
    pdf.set_x(10)
    pdf.set_fill_color(240, 240, 240)

    for key, value in config_info.items():
        value_str = "N/A" if pd.isna(value) else str(value)
        value_lines = pdf.multi_cell(value_width, line_height, value_str, border=0, split_only=True)
        row_height = line_height * len(value_lines)
        x = pdf.get_x()
        y = pdf.get_y()

        pdf.set_font("DejaVu", 'B', 11)
        pdf.multi_cell(label_width, row_height, key, border=1, fill=True)
        pdf.set_xy(x + label_width, y)
        pdf.set_font("DejaVu", '', 11)
        pdf.multi_cell(value_width, line_height, value_str, border=1)
        pdf.set_x(10)

    # ✅ Hiển thị filter mode đang sử dụng
    if filter_mode == "Task":
        filter_mode_display = "By Task"
    elif filter_mode == "Workcentre":
        filter_mode_display = "By Workcentre"
    else:
        filter_mode_display = "By Total Hours"
        
    pdf.ln(5)
    pdf.set_font("DejaVu", 'B', 11)
    pdf.cell(0, 8, f"Filter mode: {filter_mode_display}", ln=True)

    # =========================
    # 🟩 Gom biểu đồ theo project
    # =========================
    project_charts = defaultdict(list)
    for img_path, chart_title, project_name in charts_data:
        project_charts[project_name].append((img_path, chart_title))

    # =========================
    # 📄 Mỗi biểu đồ một trang
    # =========================
    for project_name, charts in project_charts.items():
        for img_path, chart_title in charts:
            if not os.path.exists(img_path):
                continue

            # ➕ Mở ảnh để xác định chiều
            img = Image.open(img_path)
            img_width, img_height = img.size
            aspect_ratio = img_height / img_width

            # Xác định chiều trang PDF
            margin = 10
            is_landscape = img_width > img_height
            orientation = 'L' if is_landscape else 'P'
            pdf.add_page(orientation=orientation)
            page_w, page_h = (297, 210) if is_landscape else (210, 297)

            # ➕ Logo và Project title
            pdf.set_font("DejaVu", 'B', 12)
            if os.path.exists(logo_path_inner):
                pdf.image(logo_path_inner, x=10, y=8, w=25)
            pdf.set_y(35)
            if project_name:
                pdf.cell(0, 6, f"Project: {project_name}", ln=True, align='C')
            else:
                pdf.cell(0, 6, "Summary Charts", ln=True, align='C')

            # ➕ Tiêu đề biểu đồ
            pdf.set_font("DejaVu", '', 11)
            pdf.ln(0.5)
            pdf.cell(0, 2, chart_title, ln=True, align='C')

            # ➕ Resize và chèn ảnh
            max_w = page_w - 2 * margin
            new_w = max_w
            new_h = new_w * aspect_ratio
            if new_h > (page_h - 2 * margin):
                new_h = page_h - 2 * margin
                new_w = new_h / aspect_ratio
            x = (page_w - new_w) / 2
            y = pdf.get_y() + 1.5
            pdf.image(img_path, x=x, y=y, w=new_w, h=new_h)

    # =========================
    # 💾 Ghi file
    # =========================
    output_dir = os.path.dirname(os.path.abspath(output_path))
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
        print(f"[DEBUG] Created output dir: {output_dir}")

    pdf.output(output_path, "F")
    return True, "✅ PDF created"

# =======================================
# CHART CREATOR (DUMMY)
# =======================================

def create_comparison_chart(df, mode, title, x_label, y_label, path, config, filter_mode="Total"):
    output_dir = "tmp_comparison"
    try:
        os.makedirs(output_dir, exist_ok=True)
        charts = {}

        df = df.copy()  # tránh cảnh báo SettingWithCopy

        # ✅ Lọc theo filter_mode
        if filter_mode == "Task":
            df = df[df['Task'] != 'All']
        elif filter_mode == "Workcentre":
            df = df[df['Workcentre'] != 'All']
        elif filter_mode == "Total":
            df.loc[:, 'Task'] = 'All'
            df.loc[:, 'Workcentre'] = 'All'

        if df.empty:
            print(f"⚠️ [DEBUG] Data trống sau lọc trong biểu đồ: mode={filter_mode}, title={title}")
            return {}

        if 'MonthName' in df.columns:
            month_order = ['January', 'February', 'March', 'April', 'May', 'June',
                           'July', 'August', 'September', 'October', 'November', 'December']
            df['MonthName'] = pd.Categorical(df['MonthName'], categories=month_order, ordered=True)

        # Biểu đồ theo thời gian (YearMonth)
        if 'Year' in df.columns and 'MonthName' in df.columns:
            df['YearMonth'] = df['Year'].astype(str) + "-" + df['MonthName'].astype(str)
            
            # ✅ Gom nhóm để tránh trùng dòng và tính tổng chính xác
            df_sorted = df.groupby(['Project Name', 'Year', 'MonthName', 'YearMonth'], as_index=False)['Total Hours'].sum()

            projects = df_sorted['Project Name'].unique()
            all_yearmonths = sorted(df_sorted['YearMonth'].unique())
            x = np.arange(len(all_yearmonths))
            width = 0.8 / len(projects) if len(projects) > 1 else 0.6

            fig, ax = plt.subplots(figsize=(15, 8.3))

            for i, project in enumerate(projects):
                df_proj = df_sorted[df_sorted['Project Name'] == project]
                y_vals = []
                for ym in all_yearmonths:
                    match = df_proj[df_proj['YearMonth'] == ym]
                    y = match['Total Hours'].sum() if not match.empty else 0  # ✅ Lấy tổng thay vì dòng đầu
                    y_vals.append(y)
                    
                ax.bar(x + i * width, y_vals, width=width, label=project)
                
                for j, val in enumerate(y_vals):
                    if val > 0:
                        ax.annotate(f"{val:.0f}", xy=(x[j] + i * width, val), xytext=(0, 5),
                                    textcoords="offset points", ha='center', fontsize=8, rotation=90)

            ax.set_title(f"{title} - Over Time")
            ax.set_xlabel(x_label)
            ax.set_ylabel(y_label)
            ax.set_xticks(x + width * (len(projects) - 1) / 2)
            ax.set_xticklabels(all_yearmonths, rotation=45, ha='right')

            ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.20), ncol=5, fontsize=8)

            plt.tight_layout()
            chart_path = os.path.join(output_dir, "chart_time.png")
            fig.savefig(chart_path, dpi=150)
            plt.close(fig)
            charts["time"] = chart_path

        # Biểu đồ theo Task
        if 'Task' in df.columns and filter_mode == "Task":
            df_task = df.groupby(['Task', 'Project Name'], as_index=False)['Total Hours'].sum()
            if df_task.empty:
                print(f"⚠️ Không có dữ liệu để vẽ biểu đồ Task cho {title}")
            else:
                df_pivot = df_task.pivot(index='Task', columns='Project Name', values='Total Hours').fillna(0)
                fig, ax = plt.subplots(figsize=(11.7, 8.3))
                bars = df_pivot.plot(kind='bar', ax=ax)
                for container in bars.containers:
                    for bar in container:
                        height = bar.get_height()
                        if height > 0:
                            ax.annotate(f"{height:.0f}", xy=(bar.get_x() + bar.get_width() / 2, height),
                                        xytext=(0, 5), textcoords="offset points", ha='center', fontsize=8, rotation=90)

                ax.set_title(f"{title} - By Task")
                ax.set_xlabel(x_label)
                ax.set_ylabel(y_label)
                ax.set_xticklabels(ax.get_xticklabels(), rotation=45, ha='right')

                ax.legend(title="Project Name", loc='upper center',
                          bbox_to_anchor=(0.5, -0.25), ncol=4, fontsize=8, frameon=False)

                plt.tight_layout()
                chart_path = os.path.join(output_dir, "chart_task.png")
                fig.savefig(chart_path, dpi=150)
                plt.close(fig)
                charts["task"] = chart_path
                
        # Biểu đồ theo Workcentre
        if 'Workcentre' in df.columns and filter_mode == "Workcentre":
            df_wc = df.groupby(['Workcentre', 'Project Name'], as_index=False)['Total Hours'].sum()
            if df_wc.empty:
                print(f"⚠️ Không có dữ liệu để vẽ biểu đồ Workcentre cho {title}")
            else:
                df_pivot = df_wc.pivot(index='Workcentre', columns='Project Name', values='Total Hours').fillna(0)
                fig, ax = plt.subplots(figsize=(15, 8.3))  # Khổ A4 ngang chuẩn

                bars = df_pivot.plot(kind='bar', ax=ax)
                # ➕ Nhãn số giờ trên cột
                for container in bars.containers:
                    for bar in container:
                        height = bar.get_height()
                        if height > 0:
                            ax.annotate(f"{height:.0f}", xy=(bar.get_x() + bar.get_width() / 2, height),
                                        xytext=(0, 5), textcoords="offset points", ha='center', fontsize=8,rotation=90)
                # ➕ Tiêu đề và trục
                ax.set_title(f"{title} - By Workcentre")
                ax.set_xlabel(x_label)
                ax.set_ylabel(y_label)
                ax.set_xticklabels(ax.get_xticklabels(), rotation=45, ha='right')

                # ✅ Legend nằm ngang bên dưới
                handles, labels = ax.get_legend_handles_labels()
                if ax.get_legend():
                    ax.get_legend().remove()
                fig.legend(
                    handles,
                    labels,
                    loc='lower center',
                    bbox_to_anchor=(0.5, -0.15),
                    ncol=min(len(labels), 5),
                    fontsize=8,
                    frameon=False
                )
                # ✅ Chừa khoảng cho legend
                fig.subplots_adjust(left=0.08, right=0.98, top=0.75, bottom=0.33)
                # ✅ BẮT BUỘC: Đảm bảo render legend trước khi lưu ảnh
                fig.canvas.draw()

                chart_path = os.path.join(output_dir, "chart_workcentre.png")
                fig.savefig(chart_path, dpi=150, bbox_inches='tight')  # ✅ THÊM bbox_inches
                plt.close(fig)
                charts["workcentre"] = chart_path
        # Biểu đồ tổng giờ (Total)
        if filter_mode == "Total":
            df_total = df.groupby("Project Name", as_index=False)["Total Hours"].sum()

            if df_total.empty:
                print("⚠️ Không có dữ liệu để vẽ biểu đồ tổng giờ theo dự án.")
                return charts

            fig, ax = plt.subplots(figsize=(15.7, 8.3))
            bars = ax.bar(df_total["Project Name"], df_total["Total Hours"])
            ax.set_title(f"{title} - Total Hours by Project")
            ax.set_xlabel(x_label)
            ax.set_ylabel(y_label)
            ax.bar_label(bars, fontsize=8, rotation=90, label_type='edge', padding=2)
            plt.xticks(rotation=45, ha='right')

            # Biểu đồ này chỉ 1 nhóm nên không cần legend
            # ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.20), ncol=5, fontsize=8)

            plt.tight_layout()
            chart_path = os.path.join(output_dir, "chart_total.png")
            fig.savefig(chart_path, dpi=150)
            plt.close(fig)
            charts["total"] = chart_path

        print(f"[DEBUG] Giá trị duy nhất trong Task: {df['Task'].unique()}")
        print(f"[DEBUG] Giá trị duy nhất trong Workcentre: {df['Workcentre'].unique()}")
        return charts

    except Exception as e:
        print(f"Chart error: {e}")
        return None

# =======================================
# EXPORT PDF COMPARISON
# =======================================

def export_comparison_pdf_report(df_comparison, comparison_config, pdf_file_path, comparison_mode, logo_path, filter_mode="Total"):
    print("=== [DEBUG] GỌI export_comparison_pdf_report ===")
    print(f"  pdf_file_path: {pdf_file_path}")
    print(f"  comparison_mode: {comparison_mode}")
    print(f"  logo_path: {logo_path}")
    print(f"  df_comparison.shape: {df_comparison.shape}")
    print(f"  comparison_config: {comparison_config}")
    print("DEBUG: df_comparison.columns =", df_comparison.columns.tolist())
    print("DEBUG: df_comparison sample:\n", df_comparison.head())
       
    if 'Hours' not in df_comparison.columns:
        raise ValueError("❌ Column 'Hours' is missing in df_comparison.")    
    if df_comparison.empty:
        print("WARNING: df_comparison is empty. Skipping PDF report export.")
        return False, "Dữ liệu rỗng"
    if not logo_path or not os.path.exists(logo_path):
        print(f"ERROR: Logo file missing or invalid: {logo_path}")
        return False, "Thiếu file logo"
    if not comparison_mode:
        return False, "❌ Thiếu chế độ so sánh (comparison_mode)"

    tmp_dir = tempfile.mkdtemp()
    try:
        success, msg = generate_comparison_pdf_report(
            df_comparison=df_comparison,
            comparison_mode=comparison_mode,
            comparison_config=comparison_config,
            pdf_file_path=pdf_file_path,
            logo_path=logo_path,
            filter_mode=filter_mode  # ✅ Truyền filter_mode xuống
        )
        return success, msg
    except Exception as e:
        return False, f"❌ Lỗi khi tạo PDF: {e}"
    finally:
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir)

# =======================================
# GENERATE PDF REPORT
# =======================================

def generate_comparison_pdf_report(df_comparison, comparison_config, pdf_file_path, comparison_mode, logo_path, filter_mode="Total"):
    tmp_dir = "tmp_comparison"
    os.makedirs(tmp_dir, exist_ok=True)
    charts_for_pdf = []

    try:
        # 🆕 Dùng project đã lọc thay vì lấy trực tiếp từ config
        filtered_projects = comparison_config.get("filtered_projects", [])  # bạn cần truyền nó từ UI vào config
        filter_mode = comparison_config.get("filter_mode", "Total")  # ✅ thêm dòng này
        
        pdf_config_info = {
            "Mode": comparison_mode,
            "Year": ', '.join(map(str, comparison_config.get('years', []))) or "N/A",
            "Months": ', '.join(comparison_config.get('months', [])) or "All",
            "Projects": ', '.join(filtered_projects) or "Không có"
        }

        # ✅ Cấu hình tiêu đề biểu đồ chung
        if comparison_mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
            chart_title = f"So sánh giờ giữa các dự án trong {comparison_config['months'][0]}, năm {comparison_config['years'][0]}"
            x_label = "Dự án"

        elif comparison_mode in ["So Sánh Dự Án Trong Một Năm", "Compare Projects in a Year"]:
            chart_title = f"So sánh giờ giữa các dự án trong năm {comparison_config['years'][0]} (theo tháng)"
            x_label = "Tháng"

        elif comparison_mode in ["So Sánh Nhiều Dự Án Qua Các Tháng/Năm", "Compare Projects Over Time (Months/Years)"]:
            chart_title = "So sánh giờ theo nhiều dự án qua các tháng và năm"
            x_label = "Năm-Tháng"
        else:
            chart_title = "Biểu đồ so sánh giờ"
            x_label = ""

        y_label = "Giờ"
        page_project_name_for_chart = "Tổng hợp nhiều dự án"

        # ✅ Gọi tạo biểu đồ tập trung (dù là mode nào cũng dùng chung)
        chart_path_placeholder = os.path.join(tmp_dir, "unused.png")  # chỉ để phù hợp với hàm gọi
        charts_dict = create_comparison_chart(
            df=df_comparison,
            mode=comparison_mode,
            title=chart_title,
            x_label=x_label,
            y_label=y_label,
            path=chart_path_placeholder,
            config=comparison_config,
            filter_mode=filter_mode # ✅ Thêm dòng này để truyền filter_mode
        )
                # ✅ Kiểm tra nếu chỉ có 1 Year-Month → bỏ biểu đồ time
        only_one_time_point = False
        if "Year" in df_comparison.columns and "MonthName" in df_comparison.columns:
            df_temp = df_comparison.copy()
            df_temp["YearMonth"] = df_temp["Year"].astype(str) + "-" + df_temp["MonthName"].astype(str)
            only_one_time_point = df_temp["YearMonth"].nunique() <= 1
        
        if charts_dict:
            print("🧪 Tổng số biểu đồ được tạo:", len(charts_dict))
            chart_title_map = {
                "time": "So sánh giờ theo thời gian",
                "total": "Tổng giờ theo từng dự án",  # ✅ thêm dòng này
                "task": "So sánh giờ theo Task giữa các dự án",
                "workcentre": "So sánh giờ theo Workcentre giữa các dự án"
            }
            print("[DEBUG] charts_dict keys:", list(charts_dict.keys()))
            
            for key in ["time", "total", "task", "workcentre"]:  # ✅ duyệt theo thứ tự ưu tiên
                if key == "time" and only_one_time_point:
                    print(f"⏭️ Bỏ qua biểu đồ '{key}' vì chỉ có 1 mốc thời gian.")
                    continue  # ✅ Bỏ qua biểu đồ time nếu không có đủ mốc thời gian
                chart_path = charts_dict.get(key)
                print(f"[DEBUG] chart {key} path = {chart_path}, exists = {os.path.exists(chart_path or '')}")
                if chart_path and os.path.exists(chart_path):
                    charts_for_pdf.append((chart_path, chart_title_map.get(key, key), page_project_name_for_chart))
        else:
            print("⚠️ charts_dict rỗng - không có biểu đồ nào được tạo từ create_comparison_chart.")
            return False, "⚠️ Không tạo được biểu đồ nào để hiển thị"
            
        if not charts_for_pdf:
            print("❌ Không có biểu đồ nào hợp lệ để tạo PDF.")
            return False, "❌ Không có biểu đồ nào tồn tại để tạo PDF"  

        # ✅ Xuất PDF
        success, msg = create_pdf_from_charts_comp(
            charts_for_pdf,
            pdf_file_path,
            "TRIAC TIME REPORT - COMPARISON",
            pdf_config_info,
            logo_path,
            filter_mode=filter_mode  # ✅ THÊM DÒNG NÀY
        )
        print(f"[DEBUG] PDF success: {success}")
        print(f"[DEBUG] PDF message: {msg}")
        print(f"[DEBUG] PDF path checked: {pdf_file_path}")
        return success, msg

    except Exception as e:
        return False, f"❌ Exception: {e}"

    finally:
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir)

def apply_comparison_filters(df_raw, comparison_config, comparison_mode, filter_mode="Total"):
    print("DEBUG: apply_comparison_filters called with:")
    if not isinstance(df_raw, pd.DataFrame):
        return pd.DataFrame(), "Dữ liệu đầu vào không hợp lệ.", []   

    print(f"  df_raw type: {type(df_raw)}")
    print(f"  comparison_config type: {type(comparison_config)}")
    print(f"  comparison_mode type: {type(comparison_mode)} value: {comparison_mode}")

    years = list(comparison_config.get('years', []))
    months = list(comparison_config.get('months', []))
    selected_projects = [p for p in comparison_config.get('selected_projects', []) if str(p).strip()]
    filter_mode = filter_mode or comparison_config.get("filter_mode", "Total")

    print("✅ Sau khi ép kiểu từ comparison_config:")
    print(f"   - Years: {years}")
    print(f"   - Months: {months}")
    print(f"   - Selected Projects: {selected_projects}")
    print(f"   - Filter Mode: {filter_mode}")

    df_filtered = df_raw.copy()
    df_filtered['Hours'] = pd.to_numeric(df_filtered['Hours'], errors='coerce').fillna(0)

    if years:
        df_filtered = df_filtered[df_filtered['Year'].isin(years)]
    if months:
        df_filtered = df_filtered[df_filtered['MonthName'].isin(months)]
        # ✅ Loại bỏ các dự án không có dữ liệu
        df_filtered_projects = df_filtered['Project name'].unique().tolist()
        selected_projects = [p for p in selected_projects if p in df_filtered_projects]
    # ✅ Luôn gán filtered_projects vào config
    comparison_config["filtered_projects"] = selected_projects
    if selected_projects:
        df_filtered = df_filtered[df_filtered['Project name'].isin(selected_projects)]
    else:
        return pd.DataFrame(), "Vui lòng chọn ít nhất một dự án để so sánh.", []
    
    if df_filtered.empty:
        return pd.DataFrame(), f"Không tìm thấy dữ liệu cho chế độ so sánh: {comparison_mode} với các lựa chọn hiện tại.", []

    title = ""

    if comparison_mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
        if len(years) != 1 or len(months) != 1 or len(selected_projects) < 2:
            return pd.DataFrame(), "Vui lòng chọn MỘT năm, MỘT tháng và ít nhất HAI dự án cho chế độ này.", []
        
        df_comparison = df_filtered.copy()
        df_comparison.rename(columns={'Project name': 'Project Name'}, inplace=True)
        df_comparison['Total Hours'] = df_comparison['Hours']
        if 'Task' not in df_comparison.columns:
            df_comparison['Task'] = 'All'
        if 'Workcentre' not in df_comparison.columns:
            df_comparison['Workcentre'] = 'All'

        # ✅ Lọc theo filter_mode nếu có
        if filter_mode == "Task":
            df_comparison = df_comparison[df_comparison["Task"].str.strip().str.lower() != "all"]
        elif filter_mode == "Workcentre":
            df_comparison = df_comparison[df_comparison["Workcentre"].str.strip().str.lower() != "all"]
        elif filter_mode == "Total":
            df_comparison['Task'] = 'All'
            df_comparison['Workcentre'] = 'All'

        title = f"So sánh giờ giữa các dự án trong {months[0]}, năm {years[0]}"
        
        print(f"✅ After filter_mode='{filter_mode}', df_comparison shape: {df_comparison.shape}")
        print(df_comparison[['Project Name', 'Task', 'Workcentre', 'Hours']].head())
        
        return df_comparison, title, selected_projects
    elif comparison_mode in ["So Sánh Dự Án Trong Một Năm", "Compare Projects in a Year"]:
        if len(years) != 1 or len(selected_projects) < 2:
            return pd.DataFrame(), "Vui lòng chọn MỘT năm và ít nhất HAI dự án cho chế độ này.", []

        df_pivot = df_filtered.groupby(['Project name', 'MonthName'])['Hours'].sum().unstack(fill_value=0)
        month_order = ['January', 'February', 'March', 'April', 'May', 'June',
                       'July', 'August', 'September', 'October', 'November', 'December']
        existing_months = [m for m in month_order if m in df_pivot.columns]
        df_pivot = df_pivot[existing_months]
        
        df_comparison = df_pivot.reset_index()
        df_comparison['Total Hours'] = df_comparison[existing_months].sum(axis=1)
        df_comparison.rename(columns={'Project name': 'Project Name'}, inplace=True)
        df_comparison['Hours'] = df_comparison['Total Hours']
        if 'Task' not in df_comparison.columns:
            df_comparison['Task'] = 'All'
        if 'Workcentre' not in df_comparison.columns:
            df_comparison['Workcentre'] = 'All'

        if filter_mode == "Task":
            df_comparison = df_comparison[df_comparison['Task'] != 'All']
        elif filter_mode == "Workcentre":
            df_comparison = df_comparison[df_comparison['Workcentre'] != 'All']
        elif filter_mode == "Total":
            df_comparison['Task'] = 'All'
            df_comparison['Workcentre'] = 'All'

        # ➕ Dòng tổng hợp
        df_total_row = pd.DataFrame([{
            'Project Name': 'Total',
            **{col: df_comparison[col].sum() for col in existing_months + ['Total Hours']}
        }])
        df_total_row['Hours'] = df_total_row['Total Hours']
        df_total_row['Task'] = 'All'
        df_total_row['Workcentre'] = 'All'

        df_comparison = pd.concat([df_comparison, df_total_row], ignore_index=True)

        title = f"So sánh giờ giữa các dự án trong năm {years[0]} (theo tháng)"
        print("📊 df_comparison shape after filter:", df_comparison.shape)
        print("📊 df_comparison preview:\n", df_comparison.head())
        return df_comparison, title, selected_projects

    elif comparison_mode in ["So Sánh Nhiều Dự Án Qua Các Tháng/Năm", "Compare Projects Over Time (Months/Years)"]:
        if not selected_projects or not years:
            return pd.DataFrame(), "Vui lòng chọn ít nhất MỘT dự án và ít nhất MỘT năm.", []

        if months:
            df_filtered = df_filtered[df_filtered['MonthName'].isin(months)]

        df_comparison = df_filtered.copy()
        df_comparison.rename(columns={'Project name': 'Project Name'}, inplace=True)
        df_comparison['Total Hours'] = df_comparison['Hours']
        if 'Task' not in df_comparison.columns:
            df_comparison['Task'] = 'All'
        if 'Workcentre' not in df_comparison.columns:
            df_comparison['Workcentre'] = 'All'

        # ✅ Lọc theo filter_mode nếu có
        if filter_mode == "Task":
            df_comparison = df_comparison[df_comparison['Task'] != 'All']
        elif filter_mode == "Workcentre":
            df_comparison = df_comparison[df_comparison['Workcentre'] != 'All']
        elif filter_mode == "Total":
            df_comparison['Task'] = 'All'
            df_comparison['Workcentre'] = 'All'

        title = "So sánh nhiều dự án qua các năm và tháng"
        return df_comparison, title, selected_projects

    return pd.DataFrame(), "❌ Chế độ so sánh không hỗ trợ.", []

def export_comparison_report(df_comparison, comparison_config, output_file_path, comparison_mode, filter_mode="Total"):
    """Xuất báo cáo so sánh ra file Excel."""
    try:
        # ✅ Đảm bảo thư mục chứa file tồn tại
        os.makedirs(os.path.dirname(output_file_path), exist_ok=True)
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            if df_comparison.empty:
                empty_df_for_excel = pd.DataFrame({"Message": ["Không có dữ liệu để hiển thị với các bộ lọc đã chọn."]})
                empty_df_for_excel.to_excel(writer, sheet_name='Comparison Report', index=False)
            else:
                df_comparison.to_excel(writer, sheet_name='Comparison Report', index=False)  

            wb = writer.book
            ws = wb['Comparison Report']

            data_last_row = ws.max_row
            info_row = data_last_row + 2 

            ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=4)
            ws.cell(row=info_row, column=1, value=f"BÁO CÁO SO SÁNH: {comparison_mode}").font = ws.cell(row=info_row, column=1).font.copy(bold=True, size=14)
            info_row += 1

            ws.cell(row=info_row, column=1, value="Năm:").font = ws.cell(row=info_row, column=1).font.copy(bold=True)
            ws.cell(row=info_row, column=2, value=', '.join(map(str, comparison_config.get('years', []))))
            info_row += 1
            ws.cell(row=info_row, column=1, value="Tháng:").font = ws.cell(row=info_row, column=1).font.copy(bold=True)
            ws.cell(row=info_row, column=2, value=', '.join(comparison_config.get('months', [])))
            info_row += 1
            ws.cell(row=info_row, column=1, value="Dự án:").font = ws.cell(row=info_row, column=1).font.copy(bold=True)
            ws.cell(row=info_row, column=2, value=', '.join(comparison_config.get('selected_projects', [])))

            if not df_comparison.empty and len(df_comparison) > 0:
                chart = None
                data_start_row = 2 
                
                df_chart_data = df_comparison.copy()
                # ✅ Lọc theo filter_mode nếu là Task hoặc Workcentre
                if filter_mode == "Task" and "Task" in df_chart_data.columns:
                    df_chart_data = df_chart_data[df_chart_data["Task"].str.strip() != "Total"]
                elif filter_mode == "Workcentre" and "Workcentre" in df_chart_data.columns:
                    df_chart_data = df_chart_data[df_chart_data["Workcentre"].str.strip() != "Total"]
                # ✅ Lọc các hàng tổng trong Project Name hoặc Year
                if 'Project Name' in df_chart_data.columns and 'Total' in df_chart_data['Project Name'].values:
                    df_chart_data = df_chart_data[df_chart_data['Project Name'] != 'Total']
                elif 'Year' in df_chart_data.columns and 'Total' in df_chart_data['Year'].values:
                    df_chart_data = df_chart_data[df_chart_data['Year'] != 'Total']
                
                if df_chart_data.empty: 
                    print("Không có đủ dữ liệu để vẽ biểu đồ so sánh sau khi loại bỏ hàng tổng.")
                    print(f"[INFO] Bỏ qua biểu đồ vì dữ liệu rỗng sau lọc (mode: {comparison_mode})")
                    wb.save(output_file_path)
                    return True

                max_row_chart = data_start_row + len(df_chart_data) - 1

                if comparison_mode in ["So Sánh Dự Án Trong Một Tháng", "Compare Projects in a Month"]:
                    chart = BarChart()
                    chart.title = "So sánh giờ theo dự án"
                    chart.x_axis.title = "Dự án"
                    chart.y_axis.title = "Giờ"
                    
                    data_ref = Reference(ws, min_col=df_comparison.columns.get_loc('Total Hours') + 1, min_row=data_start_row, max_row=max_row_chart)
                    cats_ref = Reference(ws, min_col=df_comparison.columns.get_loc('Project Name') + 1, min_row=data_start_row, max_row=max_row_chart) 
                    
                    chart.add_data(data_ref, titles_from_data=False) 
                    chart.set_categories(cats_ref)
                
                elif comparison_mode in ["So Sánh Dự Án Trong Một Năm", "Compare Projects in a Year"]:
                    chart = BarChart()
                    chart.title = "So sánh giờ theo dự án và tháng"
                    chart.x_axis.title = "Tháng"
                    chart.y_axis.title = "Giờ"
                    
                    # Cần lấy các tháng theo thứ tự đúng cho biểu đồ LineChart
                    month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
                    
                    month_cols = [col for col in df_comparison.columns if col in month_order]
                    
                    ordered_month_cols = [m for m in month_order if m in month_cols]

                    # Lấy phạm vi cho danh mục (các tháng)
                    # Giả định các tháng nằm cạnh nhau trong bảng và bắt đầu từ một cột cụ thể
                    if ordered_month_cols:
                        # openpyxl Reference uses 1-based indexing
                        ordered_month_cols = [m for m in month_order if m in month_cols]
                        min_col = df_comparison.columns.get_loc(ordered_month_cols[0]) + 1  # openpyxl 1-based
                        max_col = df_comparison.columns.get_loc(ordered_month_cols[-1]) + 1
                        
                        data_ref = Reference(ws, min_col=min_col, max_col=max_col, min_row=data_start_row, max_row=max_row_chart)
                        cats_ref = Reference(ws, min_col=min_col, min_row=1, max_col=max_col)
                        
                        chart.add_data(data_ref, titles_from_data=False)
                        chart.set_categories(cats_ref)      
                    else:
                        print("Không tìm thấy cột tháng để tạo biểu đồ.")
                        wb.save(output_file_path)
                        return True

                elif comparison_mode in ["So Sánh Nhiều Dự Án Qua Các Tháng/Năm", "Compare Projects Over Time (Months/Years)"]:

                    # Lấy tên cột chứa tổng giờ cho biểu đồ
                    total_hours_col_name = [col for col in df_comparison.columns if 'Total Hours' in col]
                    total_hours_col_name = total_hours_col_name[0] if total_hours_col_name else 'Total Hours'
                    # Tên biểu đồ tổng hợp
                    project_list = ", ".join(comparison_config.get("filtered_projects", comparison_config.get("selected_projects", [])))
                    
                    if 'MonthName' in df_comparison.columns and len(comparison_config['years']) == 1:
                    # Biểu đồ cột theo tháng
                        chart = BarChart()
                        chart.title = f"Tổng giờ các dự án ({project_list}) năm {comparison_config['years'][0]} theo tháng"
                        chart.x_axis.title = "Tháng"
                        chart.y_axis.title = "Giờ"
                        data_ref = Reference(ws, min_col=df_comparison.columns.get_loc(total_hours_col_name) + 1, min_row=data_start_row, max_row=max_row_chart)
                        cats_ref = Reference(ws, min_col=df_comparison.columns.get_loc('MonthName') + 1, min_row=data_start_row, max_row=max_row_chart)
                        chart.add_data(data_ref, titles_from_data=False)
                        chart.set_categories(cats_ref)

                    elif 'Year' in df_comparison.columns and not comparison_config['months'] and len(comparison_config['years']) > 1:
                        # Biểu đồ đường theo năm
                        chart = BarChart()
                        chart.title = f"Tổng giờ các dự án ({project_list}) theo năm"
                        chart.x_axis.title = "Năm"
                        chart.y_axis.title = "Giờ"
                        data_ref = Reference(ws, min_col=df_comparison.columns.get_loc(total_hours_col_name) + 1, min_row=data_start_row, max_row=max_row_chart)
                        cats_ref = Reference(ws, min_col=df_comparison.columns.get_loc('Year') + 1, min_row=data_start_row, max_row=max_row_chart)
                        chart.add_data(data_ref, titles_from_data=False)
                        chart.set_categories(cats_ref)
                    else:
                        raise ValueError("Không tìm thấy cấu trúc phù hợp để vẽ biểu đồ cho nhiều dự án theo tháng/năm.")

                if chart: 
                    chart_placement_row = info_row + 2
                    ws.add_chart(chart, f"A{chart_placement_row}")

            wb.save(output_file_path)
            return True
    except Exception as e:
        print(f"Lỗi khi xuất báo cáo so sánh ra Excel: {e}")
        return False

# Phần main của chương trình (có thể lấy từ main_optimized.py của bạn)
# Ví dụ cấu trúc main, bạn sẽ cần thay thế bằng nội dung thực tế của main_optimized.py
if __name__ == '__main__':
    paths = setup_paths()
    template_file = paths['template_file']
    logo_path = paths['logo_path']

    # Đảm bảo file template tồn tại
    if not os.path.exists(template_file):
        print(f"Lỗi: Không tìm thấy file template Excel '{template_file}'. Vui lòng đảm bảo file này có trong cùng thư mục với script.")
        exit()

    # Đảm bảo file logo tồn tại (nếu có)
    if not os.path.exists(logo_path):
        print(f"Cảnh báo: Không tìm thấy file logo '{logo_path}'. Báo cáo PDF sẽ được tạo mà không có logo.")
        # Nếu logo không tồn tại, bạn có thể muốn đặt logo_path thành None hoặc một đường dẫn ảnh trống
        # để tránh lỗi khi cố gắng nhúng ảnh không tồn tại.
        # Hoặc đơn giản là để hàm export_pdf_report xử lý (như hiện tại nó đã kiểm tra os.path.exists)

    raw_df = load_raw_data(template_file)
    if raw_df.empty:
        print("Không có dữ liệu thô để xử lý. Thoát chương trình.")
        exit()

    # --- Phần xử lý cho Báo cáo TIÊU CHUẨN ---
    print("\n--- Đang tạo Báo cáo TIÊU CHUẨN ---")
    standard_config = read_configs(template_file)
    standard_config['years'] = [standard_config['year']] # Chuyển year thành list cho apply_filters nếu cần
    df_standard_filtered = apply_filters(raw_df, standard_config)
    
    if not df_standard_filtered.empty:
        export_success_excel = export_report(df_standard_filtered, standard_config, paths['output_file'])
        if export_success_excel:
            print(f"Báo cáo tiêu chuẩn Excel đã được tạo thành công tại: {paths['output_file']}")
            # Tạo PDF cho báo cáo tiêu chuẩn
            export_success_pdf_standard = export_pdf_report(df_standard_filtered, standard_config, paths['pdf_report'], logo_path)
            if export_success_pdf_standard:
                print(f"Báo cáo tiêu chuẩn PDF đã được tạo thành công tại: {paths['pdf_report']}")
            else:
                print("Có lỗi khi tạo báo cáo tiêu chuẩn PDF.")
        else:
            print("Có lỗi khi tạo báo cáo tiêu chuẩn Excel.")
    else:
        print("Không có dữ liệu để tạo báo cáo tiêu chuẩn với các bộ lọc đã chọn.")

    # --- Phần xử lý cho Báo cáo SO SÁNH ---
    # Để kiểm tra chức năng so sánh, bạn cần cấu hình `Config_Year_Mode` và `Config_Project_Filter`
    # trong file `Time_report.xlsm` theo các chế độ so sánh.
    # Ví dụ minh họa cách gọi, bạn sẽ cần tùy chỉnh `comparison_config` và `comparison_mode`
    # dựa trên logic đọc cấu hình thực tế của bạn cho chế độ so sánh trong `main_optimized.py`.

    print("\n--- Đang tạo Báo cáo SO SÁNH (Ví dụ) ---")
    
    # Ví dụ cấu hình cho "So Sánh Dự Án Trong Một Tháng"
    # Bạn sẽ cần đọc cấu hình này từ file Excel của bạn theo cách tương tự `read_configs`
    # hoặc thiết lập thủ công cho mục đích thử nghiệm.
    
    # Giả định project_filter_df từ config_project_filter đã được xử lý để lấy ra các dự án được chọn
    # Trong main_optimized.py, bạn sẽ cần một logic để đọc config cho chế độ so sánh.
    # Để đơn giản trong ví dụ này, tôi sẽ giả định một cấu hình so sánh:
    
    # Lấy danh sách tất cả các Project name có trong raw_df để dùng cho việc so sánh
    all_projects_in_raw_data = raw_df['Project name'].unique().tolist()
    
    # Lấy các dự án từ config ban đầu mà có flag 'yes'
    projects_for_comparison_from_config = standard_config['project_filter_df'][
        standard_config['project_filter_df']['Include'] == 'yes'
    ]['Project Name'].tolist()
    
    if len(projects_for_comparison_from_config) >= 2 and standard_config['months']:
        # Ví dụ 1: So sánh nhiều dự án trong một tháng (nếu có đủ data và config phù hợp)
        # Sẽ cần tinh chỉnh lại để match với cấu hình đọc từ Excel
        comparison_config_month_example = {
            'years': [standard_config['year']],
            'months': [standard_config['months'][0]] if standard_config['months'] else ['January'], # Lấy tháng đầu tiên hoặc mặc định
            'selected_projects': projects_for_comparison_from_config[:2] if len(projects_for_comparison_from_config) >= 2 else all_projects_in_raw_data[:2]
        }
        comparison_config_month_example["filter_mode"] = "Total"  # ✅ gán sau khi tạo dict
        if comparison_config_month_example['selected_projects']:
            print(f"\nChế độ: So Sánh Dự Án Trong Một Tháng (năm {comparison_config_month_example['years'][0]}, tháng {comparison_config_month_example['months'][0]})")
            df_comp_month, msg_month = apply_comparison_filters(raw_df, comparison_config_month_example, "So Sánh Dự Án Trong Một Tháng")
            if not df_comp_month.empty:
                os.makedirs(os.path.dirname(paths['comparison_output_file']), exist_ok=True)
                os.makedirs(os.path.dirname(paths['comparison_pdf_report']), exist_ok=True)
                export_success_excel_comp_month = export_comparison_report(df_comp_month, comparison_config_month_example, get_comparison_excel_path("So Sánh Dự Án Trong Một Tháng", paths['comparison_output_file']), "So Sánh Dự Án Trong Một Tháng")
                if export_success_excel_comp_month:
                    print(f"Báo cáo so sánh Excel (theo tháng) đã được tạo thành công tại: {get_comparison_excel_path('So Sánh Dự Án Trong Một Tháng', paths['comparison_output_file'])}")
                    export_success_pdf_comp_month = export_comparison_pdf_report(df_comp_month, comparison_config_month_example, get_comparison_pdf_path("So Sánh Dự Án Trong Một Tháng", paths['comparison_pdf_report']), "So Sánh Dự Án Trong Một Tháng", logo_path)
                    if export_success_pdf_comp_month:
                        print(f"Báo cáo so sánh PDF (theo tháng) đã được tạo thành công tại: {get_comparison_pdf_path('So Sánh Dự Án Trong Một Tháng', paths['comparison_pdf_report'])}")
                    else:
                        print("Có lỗi khi tạo báo cáo so sánh PDF (theo tháng).")
                else:
                    print("Có lỗi khi tạo báo cáo so sánh Excel (theo tháng).")
            else:
                print(f"Không có dữ liệu cho chế độ so sánh 'So Sánh Dự Án Trong Một Tháng': {msg_month}")
        else:
            print("Không đủ dự án để thực hiện so sánh dự án trong một tháng.")

    # Ví dụ 2: So sánh một dự án qua các tháng/năm (nếu có đủ data và config phù hợp)
    if all_projects_in_raw_data:
        # Cấu hình để so sánh một dự án qua các tháng trong một năm
        if len(standard_config['months']) >= 2: # Cần ít nhất 2 tháng để so sánh
            comparison_config_single_proj_months_example = {
                'years': [standard_config['year']],
                'months': standard_config['months'],
                'selected_projects': [all_projects_in_raw_data[0]] # Chọn dự án đầu tiên
            }
            comparison_config_single_proj_months_example["filter_mode"] = "Total"
            print(f"\nChế độ: So Sánh Một Dự Án Qua Các Tháng (dự án: {comparison_config_single_proj_months_example['selected_projects'][0]}, năm {comparison_config_single_proj_months_example['years'][0]})")
            df_comp_single_proj_months, msg_single_proj_months = apply_comparison_filters(raw_df, comparison_config_single_proj_months_example, "So Sánh Một Dự Án Qua Các Tháng/Năm")
            if not df_comp_single_proj_months.empty:
                os.makedirs(os.path.dirname(paths['comparison_output_file']), exist_ok=True)
                os.makedirs(os.path.dirname(paths['comparison_pdf_report']), exist_ok=True)
                export_success_excel_comp_single_proj_months = export_comparison_report(df_comp_single_proj_months, comparison_config_single_proj_months_example, paths['comparison_output_file'].replace(".xlsx", "_SingleProjMonths.xlsx"), "So Sánh Một Dự Án Qua Các Tháng/Năm")
                if export_success_excel_comp_single_proj_months:
                    print(f"Báo cáo so sánh Excel (một dự án qua các tháng) đã được tạo thành công tại: {paths['comparison_output_file'].replace('.xlsx', '_SingleProjMonths.xlsx')}")
                    export_success_pdf_comp_single_proj_months = export_comparison_pdf_report(df_comp_single_proj_months, comparison_config_single_proj_months_example, paths['comparison_pdf_report'].replace(".pdf", "_SingleProjMonths.pdf"), "So Sánh Một Dự Án Qua Các Tháng/Năm", logo_path)
                    if export_success_pdf_comp_single_proj_months:
                        print(f"Báo cáo so sánh PDF (một dự án qua các tháng) đã được tạo thành công tại: {paths['comparison_pdf_report'].replace('.pdf', '_SingleProjMonths.pdf')}")
                    else:
                        print("Có lỗi khi tạo báo cáo so sánh PDF (một dự án qua các tháng).")
                else:
                    print("Có lỗi khi tạo báo cáo so sánh Excel (một dự án qua các tháng).")
            else:
                print(f"Không có dữ liệu cho chế độ so sánh 'So Sánh Một Dự Án Qua Các Tháng/Năm' (theo tháng): {msg_single_proj_months}")
        else:
             print("Không đủ tháng để thực hiện so sánh một dự án qua các tháng.")

        # Cấu hình để so sánh một dự án qua các năm (cần ít nhất 2 năm trong dữ liệu thô)
        # Để test phần này, raw_df cần chứa dữ liệu của nhiều năm.
        # Đảm bảo có biến path_dict để lưu các file cuối cùng
        comparison_path_dict = {}
        # Đoạn so sánh theo năm
        available_years = raw_df['Year'].unique().tolist()
        if len(available_years) >= 2:
            comparison_mode = "So Sánh Nhiều Dự Án Qua Các Tháng/Năm"
            comparison_config_single_proj_years_example = {
                'years': available_years, # Sử dụng tất cả các năm có sẵn
                'months': [], # Không lọc theo tháng
                'selected_projects': [all_projects_in_raw_data[0]] # Chọn dự án đầu tiên
            }
            comparison_config_single_proj_years_example["filter_mode"] = "Total"
            filter_mode = comparison_config_single_proj_years_example.get("filter_mode", "Total")
            print(f"\nChế độ: So Sánh Nhiều Dự Án Qua Các Tháng/Năm (dự án: {comparison_config_single_proj_years_example['selected_projects'][0]})")
            df_comp_single_proj_years, msg_single_proj_years = apply_comparison_filters(
                raw_df,
                comparison_config_single_proj_years_example,
                comparison_mode
            )
            if not df_comp_single_proj_years.empty:
                os.makedirs(os.path.dirname(paths['comparison_output_file']), exist_ok=True)
                os.makedirs(os.path.dirname(paths['comparison_pdf_report']), exist_ok=True)
                excel_path = get_comparison_excel_path(comparison_mode, paths["comparison_output_file"])
                pdf_path = get_comparison_pdf_path(comparison_mode, paths["comparison_pdf_report"])

                comparison_path_dict["comparison_output_file"] = excel_path
                comparison_path_dict["comparison_pdf_report"] = pdf_path
                
                export_success_excel = export_comparison_report(
                    df_comp_single_proj_years,
                    comparison_config_single_proj_years_example,
                    excel_path,
                    comparison_mode = "So Sánh Nhiều Dự Án Qua Các Tháng/Năm"
                )
                if export_success_excel:
                    print(f"✅ Báo cáo Excel đã tạo: {excel_path}")

                    export_success_pdf = export_comparison_pdf_report(
                        df_comp_single_proj_years,
                        comparison_config_single_proj_years_example,
                        pdf_path,
                        comparison_mode,
                        logo_path,
                        filter_mode
                    )

                    if export_success_pdf:
                        print(f"✅ Báo cáo PDF đã tạo: {pdf_path}")
                    else:
                        print("❌ Có lỗi khi tạo báo cáo PDF (một dự án qua các năm).")
                else:
                    print("❌ Có lỗi khi tạo báo cáo Excel (một dự án qua các năm).")
            else:
                print(f"⚠️ Không có dữ liệu cho '{comparison_mode}': {msg_single_proj_years}")
        else:
            print("⚠️ Không đủ năm trong dữ liệu để thực hiện so sánh một dự án qua các năm.")
